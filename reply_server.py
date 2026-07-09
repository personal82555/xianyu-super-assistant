from fastapi import FastAPI, HTTPException, Depends, status, UploadFile, File, Form, Body
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
from typing import List, Tuple, Optional, Dict, Any
from starlette.requests import Request
from pathlib import Path
from urllib.parse import unquote
import hashlib
import secrets
import time
import json
import os
import uvicorn
import pandas as pd
import io

import cookie_manager
from db_manager import db_manager
from file_log_collector import setup_file_logging, get_file_log_collector
from ai_reply_engine import ai_reply_engine
from utils.qr_login import qr_login_manager
from utils.xianyu_utils import trans_cookies
from utils.image_utils import image_manager
from utils.user_profile import get_user_profile_from_cookies
from loguru import logger

# 关键字文件路径
KEYWORDS_FILE = Path(__file__).parent / "回复关键字.txt"

# 简单的用户认证配置
ADMIN_USERNAME = "admin"
DEFAULT_ADMIN_PASSWORD = "admin123"  # 系统初始化时的默认密码
SESSION_TOKENS = {}  # 存储会话token: {token: {'user_id': int, 'username': str, 'timestamp': float}}
TOKEN_EXPIRE_TIME = 24 * 60 * 60  # token过期时间：24小时

# HTTP Bearer认证
security = HTTPBearer(auto_error=False)

# 不再需要单独的密码初始化，由数据库初始化时处理


def load_keywords() -> List[Tuple[str, str]]:
    """读取关键字→回复映射表

    文件格式支持：
        关键字<空格/制表符/冒号>回复内容
    忽略空行和以 # 开头的注释行
    """
    mapping: List[Tuple[str, str]] = []
    if not KEYWORDS_FILE.exists():
        return mapping

    with KEYWORDS_FILE.open('r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            # 尝试用\t、空格、冒号分隔
            if '\t' in line:
                key, reply = line.split('\t', 1)
            elif ' ' in line:
                key, reply = line.split(' ', 1)
            elif ':' in line:
                key, reply = line.split(':', 1)
            else:
                # 无法解析的行，跳过
                continue
            mapping.append((key.strip(), reply.strip()))
    return mapping


KEYWORDS_MAPPING = load_keywords()


# 认证相关模型
class LoginRequest(BaseModel):
    username: Optional[str] = None
    password: Optional[str] = None
    email: Optional[str] = None
    verification_code: Optional[str] = None


class LoginResponse(BaseModel):
    success: bool
    token: Optional[str] = None
    message: str
    user_id: Optional[int] = None
    username: Optional[str] = None
    is_admin: Optional[bool] = None
    captcha_required: bool = False


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str


class RegisterRequest(BaseModel):
    username: str
    email: str
    password: str
    verification_code: str


class RegisterResponse(BaseModel):
    success: bool
    message: str


class SendCodeRequest(BaseModel):
    email: str
    session_id: Optional[str] = None
    type: Optional[str] = 'register'  # 'register' 或 'login'


class SendCodeResponse(BaseModel):
    success: bool
    message: str


class CaptchaRequest(BaseModel):
    session_id: str


class CaptchaResponse(BaseModel):
    success: bool
    captcha_image: str
    session_id: str
    message: str


class VerifyCaptchaRequest(BaseModel):
    session_id: str
    captcha_code: str


class VerifyCaptchaResponse(BaseModel):
    success: bool
    message: str


def generate_token() -> str:
    """生成随机token"""
    return secrets.token_urlsafe(32)


def verify_token(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)) -> Optional[Dict[str, Any]]:
    """验证token并返回用户信息"""
    if not credentials:
        return None

    token = credentials.credentials
    if token not in SESSION_TOKENS:
        return None

    token_data = SESSION_TOKENS[token]

    # 检查token是否过期
    if time.time() - token_data['timestamp'] > TOKEN_EXPIRE_TIME:
        del SESSION_TOKENS[token]
        return None

    return token_data


def verify_admin_token(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)) -> Dict[str, Any]:
    """验证管理员token"""
    user_info = verify_token(credentials)
    if not user_info:
        raise HTTPException(status_code=401, detail="未授权访问")

    # 检查是否是管理员
    if user_info['username'] != ADMIN_USERNAME:
        raise HTTPException(status_code=403, detail="需要管理员权限")

    return user_info


def require_auth(user_info: Optional[Dict[str, Any]] = Depends(verify_token)):
    """需要认证的依赖，返回用户信息"""
    if not user_info:
        raise HTTPException(status_code=401, detail="未授权访问")
    return user_info


def get_current_user(user_info: Dict[str, Any] = Depends(require_auth)) -> Dict[str, Any]:
    """获取当前登录用户信息"""
    return user_info


def get_current_user_optional(user_info: Optional[Dict[str, Any]] = Depends(verify_token)) -> Optional[Dict[str, Any]]:
    """获取当前用户信息（可选，不强制要求登录）"""
    return user_info


def get_user_log_prefix(user_info: Dict[str, Any] = None) -> str:
    """获取用户日志前缀"""
    if user_info:
        return f"【{user_info['username']}#{user_info['user_id']}】"
    return "【系统】"


def require_admin(current_user: Dict[str, Any] = Depends(get_current_user)) -> Dict[str, Any]:
    """要求管理员权限"""
    role = current_user.get('role', 'user')
    if role != 'admin' and current_user['username'] != 'admin':
        raise HTTPException(status_code=403, detail="需要管理员权限")
    return current_user


def get_user_permissions() -> dict:
    """从系统设置中获取角色权限配置"""
    from db_manager import db_manager as _db
    try:
        val = _db.get_system_setting('user_permissions')
        if val:
            return json.loads(val)
    except Exception:
        pass
    return {}


ALL_PERMISSIONS = {
    'dashboard': '仪表盘',
    'accounts': '账号管理',
    'items': '商品管理',
    'orders': '订单管理',
    'auto_reply': '自动回复',
    'ai_conversation': 'AI对话历史',
    'items_reply': '指定商品回复',
    'cards': '卡券管理',
    'auto_delivery': '自动发货',
    'notification_channels': '通知渠道',
    'message_notifications': '消息通知',
    'item_search': '商品搜索',
    'password_settings': '密码设置',
    'system_settings': '系统设置',
    'user_management': '用户管理',
    'system_logs': '系统日志',
    'data_management': '数据管理',
    'export_backup': '备份导出',
    'import_backup': '备份导入',
}


def get_role_permissions(role: str, roles_config: dict = None) -> dict:
    """获取指定角色的权限字典"""
    if roles_config is None:
        roles_config = get_user_permissions()
    role_data = roles_config.get(role, {})
    if isinstance(role_data, dict):
        return role_data
    return {}


def has_permission(current_user: Dict[str, Any], permission: str) -> bool:
    """检查当前用户是否有指定权限"""
    username = current_user.get('username', '')
    if username == 'admin':
        return True
    role = current_user.get('role', 'user')
    roles_config = get_user_permissions()
    role_perms = get_role_permissions(role, roles_config)
    # admin 角色默认拥有所有权限
    if role == 'admin' and not role_perms:
        return True
    return role_perms.get(permission, False) is True


def require_permission(permission: str):
    """返回一个依赖项，检查用户是否有指定权限"""
    def _checker(current_user: Dict[str, Any] = Depends(get_current_user)):
        if not has_permission(current_user, permission):
            raise HTTPException(status_code=403, detail=f"没有「{ALL_PERMISSIONS.get(permission, permission)}」权限，请联系管理员")
        return current_user
    return _checker


def log_with_user(level: str, message: str, user_info: Dict[str, Any] = None):
    """带用户信息的日志记录"""
    prefix = get_user_log_prefix(user_info)
    full_message = f"{prefix} {message}"

    if level.lower() == 'info':
        logger.info(full_message)
    elif level.lower() == 'error':
        logger.error(full_message)
    elif level.lower() == 'warning':
        logger.warning(full_message)
    elif level.lower() == 'debug':
        logger.debug(full_message)
    else:
        logger.info(full_message)


def match_reply(cookie_id: str, message: str) -> Optional[str]:
    """根据 cookie_id 及消息内容匹配回复
    只有启用的账号才会匹配关键字回复
    """
    mgr = cookie_manager.manager
    if mgr is None:
        return None

    # 检查账号是否启用
    if not mgr.get_cookie_status(cookie_id):
        return None  # 禁用的账号不参与自动回复

    # 优先账号级关键字
    if mgr.get_keywords(cookie_id):
        for k, r in mgr.get_keywords(cookie_id):
            if k in message:
                return r

    # 全局关键字
    for k, r in KEYWORDS_MAPPING:
        if k in message:
            return r
    return None


class RequestModel(BaseModel):
    cookie_id: str
    msg_time: str
    user_url: str
    send_user_id: str
    send_user_name: str
    item_id: str
    send_message: str
    chat_id: str


class ResponseData(BaseModel):
    send_msg: str


class ResponseModel(BaseModel):
    code: int
    data: ResponseData


app = FastAPI(
    title="Xianyu Auto Reply API",
    version="1.0.0",
    description="闲鱼自动回复系统API",
    docs_url="/docs",
    redoc_url="/redoc"
)

# 初始化文件日志收集器
setup_file_logging()

# 添加一条测试日志
from loguru import logger
logger.info("Web服务器启动，文件日志收集器已初始化")

# 添加请求日志中间件
@app.middleware("http")
async def log_requests(request, call_next):
    start_time = time.time()

    # 获取用户信息
    user_info = "未登录"
    try:
        # 从请求头中获取Authorization
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            token = auth_header.split(" ")[1]
            if token in SESSION_TOKENS:
                token_data = SESSION_TOKENS[token]
                # 检查token是否过期
                if time.time() - token_data['timestamp'] <= TOKEN_EXPIRE_TIME:
                    user_info = f"【{token_data['username']}#{token_data['user_id']}】"
    except Exception:
        pass

    logger.info(f"🌐 {user_info} API请求: {request.method} {request.url.path}")

    response = await call_next(request)

    process_time = time.time() - start_time
    logger.info(f"✅ {user_info} API响应: {request.method} {request.url.path} - {response.status_code} ({process_time:.3f}s)")

    return response

# 提供前端静态文件
import os
static_dir = os.path.join(os.path.dirname(__file__), 'static')
if not os.path.exists(static_dir):
    os.makedirs(static_dir, exist_ok=True)

app.mount('/static', StaticFiles(directory=static_dir), name='static')

# 确保图片上传目录存在
uploads_dir = os.path.join(static_dir, 'uploads', 'images')
if not os.path.exists(uploads_dir):
    os.makedirs(uploads_dir, exist_ok=True)
    logger.info(f"创建图片上传目录: {uploads_dir}")

# 健康检查端点
@app.get('/health')
async def health_check():
    """健康检查端点，用于Docker健康检查和负载均衡器"""
    try:
        # 检查Cookie管理器状态
        manager_status = "ok" if cookie_manager.manager is not None else "error"

        # 检查数据库连接
        from db_manager import db_manager
        try:
            db_manager.get_all_cookies()
            db_status = "ok"
        except Exception:
            db_status = "error"

        # 获取系统状态
        import psutil
        cpu_percent = psutil.cpu_percent(interval=1)
        memory_info = psutil.virtual_memory()

        status = {
            "status": "healthy" if manager_status == "ok" and db_status == "ok" else "unhealthy",
            "timestamp": time.time(),
            "services": {
                "cookie_manager": manager_status,
                "database": db_status
            },
            "system": {
                "cpu_percent": cpu_percent,
                "memory_percent": memory_info.percent,
                "memory_available": memory_info.available
            }
        }

        if status["status"] == "unhealthy":
            raise HTTPException(status_code=503, detail=status)

        return status

    except Exception as e:
        return {
            "status": "unhealthy",
            "timestamp": time.time(),
            "error": str(e)
        }


# 重定向根路径到登录页面
@app.get('/', response_class=HTMLResponse)
async def root():
    login_path = os.path.join(static_dir, 'login.html')
    if os.path.exists(login_path):
        with open(login_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(f.read())
    else:
        return HTMLResponse('<h3>Login page not found</h3>')


# 登录页面路由
@app.get('/login.html', response_class=HTMLResponse)
async def login_page():
    login_path = os.path.join(static_dir, 'login.html')
    if os.path.exists(login_path):
        with open(login_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(f.read())
    else:
        return HTMLResponse('<h3>Login page not found</h3>')


# 注册页面路由
@app.get('/register.html', response_class=HTMLResponse)
async def register_page():
    # 检查注册是否开启
    from db_manager import db_manager
    registration_enabled = db_manager.get_system_setting('registration_enabled')
    if registration_enabled != 'true':
        return HTMLResponse('''
        <!DOCTYPE html>
        <html>
        <head>
            <title>注册已关闭</title>
            <meta charset="utf-8">
            <style>
                body { font-family: Arial, sans-serif; text-align: center; padding: 50px; }
                .message { color: #666; font-size: 18px; }
                .back-link { margin-top: 20px; }
                .back-link a { color: #007bff; text-decoration: none; }
            </style>
        </head>
        <body>
            <h2>🚫 注册功能已关闭</h2>
            <p class="message">系统管理员已关闭用户注册功能</p>
            <div class="back-link">
                <a href="/">← 返回首页</a>
            </div>
        </body>
        </html>
        ''', status_code=403)

    register_path = os.path.join(static_dir, 'register.html')
    if os.path.exists(register_path):
        with open(register_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(f.read())
    else:
        return HTMLResponse('<h3>Register page not found</h3>')


# 管理页面（不需要服务器端认证，由前端JavaScript处理）
@app.get('/admin', response_class=HTMLResponse)
async def admin_page():
    index_path = os.path.join(static_dir, 'index.html')
    if not os.path.exists(index_path):
        return HTMLResponse('<h3>No front-end found</h3>')
    with open(index_path, 'r', encoding='utf-8') as f:
        return HTMLResponse(f.read())


# 用户管理页面路由
@app.get('/user_management.html', response_class=HTMLResponse)
async def user_management_page():
    page_path = os.path.join(static_dir, 'user_management.html')
    if os.path.exists(page_path):
        with open(page_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(f.read())
    else:
        return HTMLResponse('<h3>User management page not found</h3>')


# 日志管理页面路由
@app.get('/log_management.html', response_class=HTMLResponse)
async def log_management_page():
    page_path = os.path.join(static_dir, 'log_management.html')
    if os.path.exists(page_path):
        with open(page_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(f.read())
    else:
        return HTMLResponse('<h3>Log management page not found</h3>')


# 数据管理页面路由
@app.get('/data_management.html', response_class=HTMLResponse)
async def data_management_page():
    page_path = os.path.join(static_dir, 'data_management.html')
    if os.path.exists(page_path):
        with open(page_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(f.read())
    else:
        return HTMLResponse('<h3>Data management page not found</h3>')





# 商品搜索页面路由
@app.get('/item_search.html', response_class=HTMLResponse)
async def item_search_page():
    page_path = os.path.join(static_dir, 'item_search.html')
    if os.path.exists(page_path):
        with open(page_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(f.read())
    else:
        return HTMLResponse('<h3>Item search page not found</h3>')


# 登录接口
@app.post('/login')
async def login(request: LoginRequest):
    from db_manager import db_manager

    # 判断登录方式
    if request.username and request.password:
        # 用户名/密码登录
        logger.info(f"【{request.username}】尝试用户名登录")

        # 统一使用用户表验证（包括admin用户）
        if db_manager.verify_user_password(request.username, request.password):
            user = db_manager.get_user_by_username(request.username)
            if user:
                # 生成token
                token = generate_token()
                SESSION_TOKENS[token] = {
                    'user_id': user['id'],
                    'username': user['username'],
                    'role': user.get('role', 'user'),
                    'timestamp': time.time()
                }

                # 区分管理员和普通用户的日志
                if user['username'] == ADMIN_USERNAME:
                    logger.info(f"【{user['username']}#{user['id']}】登录成功（管理员）")
                else:
                    logger.info(f"【{user['username']}#{user['id']}】登录成功")

                return LoginResponse(
                    success=True,
                    token=token,
                    message="登录成功",
                    user_id=user['id'],
                    username=user['username'],
                    is_admin=(user['username'] == ADMIN_USERNAME)
                )

        logger.warning(f"【{request.username}】登录失败：用户名或密码错误")
        return LoginResponse(
            success=False,
            message="用户名或密码错误",
            captcha_required=True
        )

    elif request.email and request.password:
        # 邮箱/密码登录
        logger.info(f"【{request.email}】尝试邮箱密码登录")

        user = db_manager.get_user_by_email(request.email)
        if user and db_manager.verify_user_password(user['username'], request.password):
            # 生成token
            token = generate_token()
            SESSION_TOKENS[token] = {
                'user_id': user['id'],
                'username': user['username'],
                'role': user.get('role', 'user'),
                'timestamp': time.time()
            }

            logger.info(f"【{user['username']}#{user['id']}】邮箱登录成功")

            return LoginResponse(
                success=True,
                token=token,
                message="登录成功",
                user_id=user['id'],
                username=user['username'],
                is_admin=(user['username'] == ADMIN_USERNAME)
            )

        logger.warning(f"【{request.email}】邮箱登录失败：邮箱或密码错误")
        return LoginResponse(
            success=False,
            message="邮箱或密码错误",
            captcha_required=True
        )

    elif request.email and request.verification_code:
        # 邮箱/验证码登录
        logger.info(f"【{request.email}】尝试邮箱验证码登录")

        # 验证邮箱验证码
        if not db_manager.verify_email_code(request.email, request.verification_code, 'login'):
            logger.warning(f"【{request.email}】验证码登录失败：验证码错误或已过期")
            return LoginResponse(
                success=False,
                message="验证码错误或已过期"
            )

        # 获取用户信息
        user = db_manager.get_user_by_email(request.email)
        if not user:
            logger.warning(f"【{request.email}】验证码登录失败：用户不存在")
            return LoginResponse(
                success=False,
                message="用户不存在"
            )

        # 生成token
        token = generate_token()
        SESSION_TOKENS[token] = {
            'user_id': user['id'],
            'username': user['username'],
            'role': user.get('role', 'user'),
            'timestamp': time.time()
        }

        logger.info(f"【{user['username']}#{user['id']}】验证码登录成功")

        return LoginResponse(
            success=True,
            token=token,
            message="登录成功",
            user_id=user['id'],
            username=user['username'],
            is_admin=(user['username'] == ADMIN_USERNAME)
        )

    else:
        return LoginResponse(
            success=False,
            message="请提供有效的登录信息"
        )


# 验证token接口
@app.get('/verify')
async def verify(user_info: Optional[Dict[str, Any]] = Depends(verify_token)):
    if user_info:
        role = user_info.get('role', 'user')
        return {
            "authenticated": True,
            "user_id": user_info['user_id'],
            "username": user_info['username'],
            "role": role,
            "is_admin": role == 'admin' or user_info['username'] == ADMIN_USERNAME
        }
    return {"authenticated": False}


# 登出接口
@app.post('/logout')
async def logout(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)):
    if credentials and credentials.credentials in SESSION_TOKENS:
        del SESSION_TOKENS[credentials.credentials]
    return {"message": "已登出"}


# 修改管理员密码接口
@app.post('/change-admin-password')
async def change_admin_password(request: ChangePasswordRequest, admin_user: Dict[str, Any] = Depends(verify_admin_token)):
    from db_manager import db_manager

    try:
        # 验证当前密码（使用用户表验证）
        if not db_manager.verify_user_password('admin', request.current_password):
            return {"success": False, "message": "当前密码错误"}

        # 更新密码（使用用户表更新）
        success = db_manager.update_user_password('admin', request.new_password)

        if success:
            logger.info(f"【admin#{admin_user['user_id']}】管理员密码修改成功")
            return {"success": True, "message": "密码修改成功"}
        else:
            return {"success": False, "message": "密码修改失败"}

    except Exception as e:
        logger.error(f"修改管理员密码异常: {e}")
        return {"success": False, "message": "系统错误"}


# 生成图形验证码接口
@app.post('/generate-captcha')
async def generate_captcha(request: CaptchaRequest):
    from db_manager import db_manager

    try:
        # 生成图形验证码
        captcha_text, captcha_image = db_manager.generate_captcha()

        if not captcha_image:
            return CaptchaResponse(
                success=False,
                captcha_image="",
                session_id=request.session_id,
                message="图形验证码生成失败"
            )

        # 保存验证码到数据库
        if db_manager.save_captcha(request.session_id, captcha_text):
            return CaptchaResponse(
                success=True,
                captcha_image=captcha_image,
                session_id=request.session_id,
                message="图形验证码生成成功"
            )
        else:
            return CaptchaResponse(
                success=False,
                captcha_image="",
                session_id=request.session_id,
                message="图形验证码保存失败"
            )

    except Exception as e:
        logger.error(f"生成图形验证码失败: {e}")
        return CaptchaResponse(
            success=False,
            captcha_image="",
            session_id=request.session_id,
            message="图形验证码生成失败"
        )


# 验证图形验证码接口
@app.post('/verify-captcha')
async def verify_captcha(request: VerifyCaptchaRequest):
    from db_manager import db_manager

    try:
        if db_manager.verify_captcha(request.session_id, request.captcha_code):
            return VerifyCaptchaResponse(
                success=True,
                message="图形验证码验证成功"
            )
        else:
            return VerifyCaptchaResponse(
                success=False,
                message="图形验证码错误或已过期"
            )

    except Exception as e:
        logger.error(f"验证图形验证码失败: {e}")
        return VerifyCaptchaResponse(
            success=False,
            message="图形验证码验证失败"
        )


# 发送验证码接口（需要先验证图形验证码）
@app.post('/send-verification-code')
async def send_verification_code(request: SendCodeRequest):
    from db_manager import db_manager

    try:
        # 检查是否已验证图形验证码
        # 通过检查数据库中是否存在已验证的图形验证码记录
        with db_manager.lock:
            cursor = db_manager.conn.cursor()
            current_time = time.time()

            # 查找最近5分钟内该session_id的验证记录
            # 由于验证成功后验证码会被删除，我们需要另一种方式来跟踪验证状态
            # 这里我们检查该session_id是否在最近验证过（通过检查是否有已删除的记录）

            # 为了简化，我们要求前端在验证图形验证码成功后立即发送邮件验证码
            # 或者我们可以在验证成功后设置一个临时标记
            pass

        # 根据验证码类型进行不同的检查
        if request.type == 'register':
            # 注册验证码：检查邮箱是否已注册
            existing_user = db_manager.get_user_by_email(request.email)
            if existing_user:
                return SendCodeResponse(
                    success=False,
                    message="该邮箱已被注册"
                )
        elif request.type == 'login':
            # 登录验证码：检查邮箱是否存在
            existing_user = db_manager.get_user_by_email(request.email)
            if not existing_user:
                return SendCodeResponse(
                    success=False,
                    message="该邮箱未注册"
                )

        # 生成验证码
        code = db_manager.generate_verification_code()

        # 保存验证码到数据库
        if not db_manager.save_verification_code(request.email, code, request.type):
            return SendCodeResponse(
                success=False,
                message="验证码保存失败，请稍后重试"
            )

        # 发送验证码邮件
        if await db_manager.send_verification_email(request.email, code):
            return SendCodeResponse(
                success=True,
                message="验证码已发送到您的邮箱，请查收"
            )
        else:
            return SendCodeResponse(
                success=False,
                message="验证码发送失败，请检查邮箱地址或稍后重试"
            )

    except Exception as e:
        logger.error(f"发送验证码失败: {e}")
        return SendCodeResponse(
            success=False,
            message="发送验证码失败，请稍后重试"
        )


# 用户注册接口
@app.post('/register')
async def register(request: RegisterRequest):
    from db_manager import db_manager

    # 检查注册是否开启
    registration_enabled = db_manager.get_system_setting('registration_enabled')
    if registration_enabled != 'true':
        logger.warning(f"【{request.username}】注册失败: 注册功能已关闭")
        return RegisterResponse(
            success=False,
            message="注册功能已关闭，请联系管理员"
        )

    try:
        logger.info(f"【{request.username}】尝试注册，邮箱: {request.email}")

        # 验证邮箱验证码
        if not db_manager.verify_email_code(request.email, request.verification_code):
            logger.warning(f"【{request.username}】注册失败: 验证码错误或已过期")
            return RegisterResponse(
                success=False,
                message="验证码错误或已过期"
            )

        # 检查用户名是否已存在
        existing_user = db_manager.get_user_by_username(request.username)
        if existing_user:
            logger.warning(f"【{request.username}】注册失败: 用户名已存在")
            return RegisterResponse(
                success=False,
                message="用户名已存在"
            )

        # 检查邮箱是否已注册
        existing_email = db_manager.get_user_by_email(request.email)
        if existing_email:
            logger.warning(f"【{request.username}】注册失败: 邮箱已被注册")
            return RegisterResponse(
                success=False,
                message="该邮箱已被注册"
            )

        # 创建用户
        if db_manager.create_user(request.username, request.email, request.password):
            logger.info(f"【{request.username}】注册成功")
            return RegisterResponse(
                success=True,
                message="注册成功，请登录"
            )
        else:
            logger.error(f"【{request.username}】注册失败: 数据库操作失败")
            return RegisterResponse(
                success=False,
                message="注册失败，请稍后重试"
            )

    except Exception as e:
        logger.error(f"【{request.username}】注册异常: {e}")
        return RegisterResponse(
            success=False,
            message="注册失败，请稍后重试"
        )


@app.post("/xianyu/reply", response_model=ResponseModel)
async def xianyu_reply(req: RequestModel):
    msg_template = match_reply(req.cookie_id, req.send_message)
    is_default_reply = False

    if not msg_template:
        # 从数据库获取默认回复
        from db_manager import db_manager
        default_reply_settings = db_manager.get_default_reply(req.cookie_id)

        if default_reply_settings and default_reply_settings.get('enabled', False):
            # 检查是否开启了"只回复一次"功能
            if default_reply_settings.get('reply_once', False):
                # 检查是否已经回复过这个chat_id
                if db_manager.has_default_reply_record(req.cookie_id, req.chat_id):
                    raise HTTPException(status_code=404, detail="该对话已使用默认回复，不再重复回复")

            msg_template = default_reply_settings.get('reply_content', '')
            is_default_reply = True

        # 如果数据库中没有设置或为空，返回错误
        if not msg_template:
            raise HTTPException(status_code=404, detail="未找到匹配的回复规则且未设置默认回复")

    # 按占位符格式化
    try:
        send_msg = msg_template.format(
            send_user_id=req.send_user_id,
            send_user_name=req.send_user_name,
            send_message=req.send_message,
        )
    except Exception:
        # 如果格式化失败，返回原始内容
        send_msg = msg_template

    # 如果是默认回复且开启了"只回复一次"，记录回复记录
    if is_default_reply:
        from db_manager import db_manager
        default_reply_settings = db_manager.get_default_reply(req.cookie_id)
        if default_reply_settings and default_reply_settings.get('reply_once', False):
            db_manager.add_default_reply_record(req.cookie_id, req.chat_id)

    return {"code": 200, "data": {"send_msg": send_msg}}

# ------------------------- 账号 / 关键字管理接口 -------------------------


class CookieIn(BaseModel):
    id: str
    value: str


class CookieStatusIn(BaseModel):
    enabled: bool


class DefaultReplyIn(BaseModel):
    enabled: bool
    reply_content: Optional[str] = None
    reply_once: bool = False


class NotificationChannelIn(BaseModel):
    name: str
    type: str = "qq"
    config: str


class NotificationChannelUpdate(BaseModel):
    name: str
    config: str
    enabled: bool = True


class MessageNotificationIn(BaseModel):
    channel_id: int
    enabled: bool = True


class SystemSettingIn(BaseModel):
    key: str
    value: str
    description: Optional[str] = None





@app.get("/cookies")
def list_cookies(current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        return []

    # 获取当前用户的cookies
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)
    return list(user_cookies.keys())


@app.get("/cookies/details")
def get_cookies_details(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取所有Cookie的详细信息（包括值和状态）"""
    if cookie_manager.manager is None:
        return []

    # 获取当前用户的cookies
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    result = []
    for cookie_id, cookie_value in user_cookies.items():
        cookie_enabled = cookie_manager.manager.get_cookie_status(cookie_id)
        auto_confirm = db_manager.get_auto_confirm(cookie_id)
        # 获取备注信息
        cookie_details = db_manager.get_cookie_details(cookie_id)
        remark = cookie_details.get('remark', '') if cookie_details else ''

        created_at = cookie_details.get('created_at') if cookie_details else None
        nickname = cookie_details.get('nickname', '') if cookie_details else ''
        avatar_url = cookie_details.get('avatar_url', '') if cookie_details else ''
        
        result.append({
            'id': cookie_id,
            'value': cookie_value,
            'enabled': cookie_enabled,
            'auto_confirm': auto_confirm,
            'remark': remark,
            'pause_duration': cookie_details.get('pause_duration', 10) if cookie_details else 10,
            'created_at': created_at,
            'nickname': nickname,
            'avatar_url': avatar_url
        })
    return result


@app.post("/cookies")
def add_cookie(item: CookieIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 添加cookie时绑定到当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager

        log_with_user('info', f"尝试添加Cookie: {item.id}, 当前用户ID: {user_id}, 用户名: {current_user.get('username', 'unknown')}", current_user)

        # 检查cookie是否已存在且属于其他用户
        existing_cookies = db_manager.get_all_cookies()
        if item.id in existing_cookies:
            # 检查是否属于当前用户
            user_cookies = db_manager.get_all_cookies(user_id)
            if item.id not in user_cookies:
                log_with_user('warning', f"Cookie ID冲突: {item.id} 已被其他用户使用", current_user)
                raise HTTPException(status_code=400, detail="该Cookie ID已被其他用户使用")

        # 保存到数据库时指定用户ID
        db_manager.save_cookie(item.id, item.value, user_id)

        # 添加到CookieManager，同时指定用户ID
        cookie_manager.manager.add_cookie(item.id, item.value, user_id=user_id)
        log_with_user('info', f"Cookie添加成功: {item.id}", current_user)
        return {"msg": "success"}
    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"添加Cookie失败: {item.id} - {str(e)}", current_user)
        raise HTTPException(status_code=400, detail=str(e))


@app.put('/cookies/{cid}')
def update_cookie(cid: str, item: CookieIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail='CookieManager 未就绪')
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 更新cookie时保持用户绑定
        db_manager.save_cookie(cid, item.value, user_id)
        cookie_manager.manager.update_cookie(cid, item.value)
        return {'msg': 'updated'}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# ========================= 扫码登录相关接口 =========================

@app.post("/qr-login/generate")
async def generate_qr_code(current_user: Dict[str, Any] = Depends(get_current_user)):
    """生成扫码登录二维码"""
    try:
        log_with_user('info', "请求生成扫码登录二维码", current_user)

        result = await qr_login_manager.generate_qr_code()

        if result['success']:
            log_with_user('info', f"扫码登录二维码生成成功: {result['session_id']}", current_user)
        else:
            log_with_user('warning', f"扫码登录二维码生成失败: {result.get('message', '未知错误')}", current_user)

        return result

    except Exception as e:
        log_with_user('error', f"生成扫码登录二维码异常: {str(e)}", current_user)
        return {'success': False, 'message': f'生成二维码失败: {str(e)}'}


@app.get("/qr-login/check/{session_id}")
async def check_qr_code_status(session_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """检查扫码登录状态"""
    try:
        # 清理过期会话
        qr_login_manager.cleanup_expired_sessions()

        # 获取会话状态
        status_info = qr_login_manager.get_session_status(session_id)

        if status_info['status'] == 'success':
            # 登录成功，处理Cookie
            cookies_info = qr_login_manager.get_session_cookies(session_id)
            if cookies_info:
                account_info = await process_qr_login_cookies(
                    cookies_info['cookies'],
                    cookies_info['unb'],
                    current_user
                )
                status_info['account_info'] = account_info

                log_with_user('info', f"扫码登录成功处理完成: {session_id}, 账号: {account_info.get('account_id', 'unknown')}", current_user)

        return status_info

    except Exception as e:
        log_with_user('error', f"检查扫码登录状态异常: {str(e)}", current_user)
        return {'status': 'error', 'message': str(e)}


async def process_qr_login_cookies(cookies: str, unb: str, current_user: Dict[str, Any]) -> Dict[str, Any]:
    """处理扫码登录获取的Cookie"""
    try:
        user_id = current_user['user_id']

        # 获取用户资料（昵称和头像）
        nickname = ''
        avatar_url = ''
        try:
            user_profile = await get_user_profile_from_cookies(cookies)
            if user_profile:
                nickname = user_profile.get('nickname', '')
                avatar_url = user_profile.get('avatar_url', '')
                log_with_user('info', f"获取用户资料成功: nickname={nickname}, avatar_url={avatar_url[:50]}..." if avatar_url else f"获取用户资料成功: nickname={nickname}", current_user)
            else:
                log_with_user('warning', "获取用户资料失败，将使用默认值", current_user)
        except Exception as e:
            log_with_user('warning', f"获取用户资料异常: {e}", current_user)

        # 检查是否已存在相同unb的账号
        existing_cookies = db_manager.get_all_cookies(user_id)
        existing_account_id = None

        for account_id, cookie_value in existing_cookies.items():
            try:
                # 解析现有Cookie中的unb
                existing_cookie_dict = trans_cookies(cookie_value)
                if existing_cookie_dict.get('unb') == unb:
                    existing_account_id = account_id
                    break
            except:
                continue

        if existing_account_id:
            # 更新现有账号的Cookie
            db_manager.save_cookie(existing_account_id, cookies, user_id)
            
            # 更新用户资料
            if nickname or avatar_url:
                db_manager.update_cookie_user_info(existing_account_id, nickname, avatar_url)

            # 更新cookie_manager中的Cookie
            if cookie_manager.manager:
                cookie_manager.manager.update_cookie(existing_account_id, cookies)

            log_with_user('info', f"扫码登录更新现有账号Cookie: {existing_account_id}, UNB: {unb}, nickname: {nickname}", current_user)

            return {
                'account_id': existing_account_id,
                'is_new_account': False,
                'nickname': nickname,
                'avatar_url': avatar_url
            }
        else:
            # 创建新账号，使用unb作为账号ID
            account_id = unb

            # 确保账号ID唯一
            counter = 1
            original_account_id = account_id
            while account_id in existing_cookies:
                account_id = f"{original_account_id}_{counter}"
                counter += 1

            # 保存新账号
            db_manager.save_cookie(account_id, cookies, user_id)
            
            # 保存用户资料
            if nickname or avatar_url:
                db_manager.update_cookie_user_info(account_id, nickname, avatar_url)

            # 添加到cookie_manager
            if cookie_manager.manager:
                cookie_manager.manager.add_cookie(account_id, cookies)

            log_with_user('info', f"扫码登录创建新账号: {account_id}, UNB: {unb}, nickname: {nickname}", current_user)

            return {
                'account_id': account_id,
                'is_new_account': True,
                'nickname': nickname,
                'avatar_url': avatar_url
            }

    except Exception as e:
        log_with_user('error', f"处理扫码登录Cookie失败: {str(e)}", current_user)
        raise e


@app.put('/cookies/{cid}/status')
def update_cookie_status(cid: str, status_data: CookieStatusIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号的启用/禁用状态"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail='CookieManager 未就绪')
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        cookie_manager.manager.update_cookie_status(cid, status_data.enabled)
        return {'msg': 'status updated', 'enabled': status_data.enabled}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# ------------------------- 默认回复管理接口 -------------------------

@app.get('/default-replies/{cid}')
def get_default_reply(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定账号的默认回复设置"""
    from db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        result = db_manager.get_default_reply(cid)
        if result is None:
            # 如果没有设置，返回默认值
            return {'enabled': False, 'reply_content': '', 'reply_once': False}
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put('/default-replies/{cid}')
def update_default_reply(cid: str, reply_data: DefaultReplyIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新指定账号的默认回复设置"""
    from db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        db_manager.save_default_reply(cid, reply_data.enabled, reply_data.reply_content, reply_data.reply_once)
        return {'msg': 'default reply updated', 'enabled': reply_data.enabled, 'reply_once': reply_data.reply_once}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get('/default-replies')
def get_all_default_replies(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户所有账号的默认回复设置"""
    from db_manager import db_manager
    try:
        # 只返回当前用户的默认回复设置
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        all_replies = db_manager.get_all_default_replies()
        # 过滤只属于当前用户的回复设置
        user_replies = {cid: reply for cid, reply in all_replies.items() if cid in user_cookies}
        return user_replies
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete('/default-replies/{cid}')
def delete_default_reply(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """删除指定账号的默认回复设置"""
    from db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        success = db_manager.delete_default_reply(cid)
        if success:
            return {'msg': 'default reply deleted'}
        else:
            raise HTTPException(status_code=400, detail='删除失败')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post('/default-replies/{cid}/clear-records')
def clear_default_reply_records(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """清空指定账号的默认回复记录"""
    from db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        db_manager.clear_default_reply_records(cid)
        return {'msg': 'default reply records cleared'}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 通知渠道管理接口 -------------------------

@app.get('/notification-channels')
def get_notification_channels(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取所有通知渠道"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        return db_manager.get_notification_channels(user_id)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post('/notification-channels')
def create_notification_channel(channel_data: NotificationChannelIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """创建通知渠道"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        channel_id = db_manager.create_notification_channel(
            channel_data.name,
            channel_data.type,
            channel_data.config,
            user_id
        )
        return {'msg': 'notification channel created', 'id': channel_id}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get('/notification-channels/{channel_id}')
def get_notification_channel(channel_id: int, _: None = Depends(require_auth)):
    """获取指定通知渠道"""
    from db_manager import db_manager
    try:
        channel = db_manager.get_notification_channel(channel_id)
        if not channel:
            raise HTTPException(status_code=404, detail='通知渠道不存在')
        return channel
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put('/notification-channels/{channel_id}')
def update_notification_channel(channel_id: int, channel_data: NotificationChannelUpdate, _: None = Depends(require_auth)):
    """更新通知渠道"""
    from db_manager import db_manager
    try:
        success = db_manager.update_notification_channel(
            channel_id,
            channel_data.name,
            channel_data.config,
            channel_data.enabled
        )
        if success:
            return {'msg': 'notification channel updated'}
        else:
            raise HTTPException(status_code=404, detail='通知渠道不存在')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.delete('/notification-channels/{channel_id}')
def delete_notification_channel(channel_id: int, _: None = Depends(require_auth)):
    """删除通知渠道"""
    from db_manager import db_manager
    try:
        success = db_manager.delete_notification_channel(channel_id)
        if success:
            return {'msg': 'notification channel deleted'}
        else:
            raise HTTPException(status_code=404, detail='通知渠道不存在')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 消息通知配置接口 -------------------------

@app.get('/message-notifications')
def get_all_message_notifications(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户所有账号的消息通知配置"""
    from db_manager import db_manager
    try:
        # 只返回当前用户的消息通知配置
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        all_notifications = db_manager.get_all_message_notifications()
        # 过滤只属于当前用户的通知配置
        user_notifications = {cid: notifications for cid, notifications in all_notifications.items() if cid in user_cookies}
        return user_notifications
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get('/message-notifications/{cid}')
def get_account_notifications(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定账号的消息通知配置"""
    from db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        return db_manager.get_account_notifications(cid)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post('/message-notifications/{cid}')
def set_message_notification(cid: str, notification_data: MessageNotificationIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """设置账号的消息通知"""
    from db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 检查通知渠道是否存在
        channel = db_manager.get_notification_channel(notification_data.channel_id)
        if not channel:
            raise HTTPException(status_code=404, detail='通知渠道不存在')

        success = db_manager.set_message_notification(cid, notification_data.channel_id, notification_data.enabled)
        if success:
            return {'msg': 'message notification set'}
        else:
            raise HTTPException(status_code=400, detail='设置失败')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete('/message-notifications/account/{cid}')
def delete_account_notifications(cid: str, _: None = Depends(require_auth)):
    """删除账号的所有消息通知配置"""
    from db_manager import db_manager
    try:
        success = db_manager.delete_account_notifications(cid)
        if success:
            return {'msg': 'account notifications deleted'}
        else:
            raise HTTPException(status_code=404, detail='账号通知配置不存在')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete('/message-notifications/{notification_id}')
def delete_message_notification(notification_id: int, _: None = Depends(require_auth)):
    """删除消息通知配置"""
    from db_manager import db_manager
    try:
        success = db_manager.delete_message_notification(notification_id)
        if success:
            return {'msg': 'message notification deleted'}
        else:
            raise HTTPException(status_code=404, detail='通知配置不存在')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 系统设置接口 -------------------------

@app.get('/system-settings')
def get_system_settings(_: None = Depends(require_auth)):
    """获取系统设置（排除敏感信息）"""
    from db_manager import db_manager
    try:
        settings = db_manager.get_all_system_settings()
        # 移除敏感信息
        if 'admin_password_hash' in settings:
            del settings['admin_password_hash']
        return settings
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))





@app.put('/system-settings/ai-defaults')
async def update_ai_defaults(request: Request, _: None = Depends(require_auth)):
    """批量保存AI回复全局默认设置"""
    from db_manager import db_manager
    try:
        settings = await request.json()
        mapping = {
            'ai_default_enabled': 'ai_default_enabled',
            'ai_default_model': 'ai_default_model',
            'ai_default_base_url': 'ai_default_base_url',
            'ai_default_api_key': 'ai_default_api_key',
            'ai_default_max_discount_percent': 'ai_default_max_discount_percent',
            'ai_default_max_discount_amount': 'ai_default_max_discount_amount',
            'ai_default_max_bargain_rounds': 'ai_default_max_bargain_rounds',
            'ai_default_prompts': 'ai_default_prompts',
            'ai_default_refund_policy': 'ai_default_refund_policy',
        }
        for key, db_key in mapping.items():
            if key in settings:
                db_manager.set_system_setting(db_key, str(settings[key]), f'AI回复-{key}')
        # 清理AI引擎缓存，使新设置生效
        from ai_reply_engine import ai_reply_engine
        ai_reply_engine.clear_client_cache()
        return {'msg': 'AI global defaults updated'}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put('/system-settings/{key}')
def update_system_setting(key: str, setting_data: SystemSettingIn, _: None = Depends(require_auth)):
    """更新系统设置"""
    from db_manager import db_manager
    try:
        # 禁止直接修改密码哈希
        if key == 'admin_password_hash':
            raise HTTPException(status_code=400, detail='请使用密码修改接口')

        success = db_manager.set_system_setting(key, setting_data.value, setting_data.description)
        if success:
            return {'msg': 'system setting updated'}
        else:
            raise HTTPException(status_code=400, detail='更新失败')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 注册设置接口 -------------------------

@app.get('/registration-status')
def get_registration_status():
    """获取注册开关状态（公开接口，无需认证）"""
    from db_manager import db_manager
    try:
        enabled_str = db_manager.get_system_setting('registration_enabled')
        logger.info(f"从数据库获取的注册设置值: '{enabled_str}'")  # 调试信息

        # 如果设置不存在，默认为开启
        if enabled_str is None:
            enabled_bool = True
            message = '注册功能已开启'
        else:
            enabled_bool = enabled_str == 'true'
            message = '注册功能已开启' if enabled_bool else '注册功能已关闭'

        logger.info(f"解析后的注册状态: enabled={enabled_bool}, message='{message}'")  # 调试信息

        return {
            'enabled': enabled_bool,
            'message': message
        }
    except Exception as e:
        logger.error(f"获取注册状态失败: {e}")
        return {'enabled': True, 'message': '注册功能已开启'}  # 出错时默认开启


class RegistrationSettingUpdate(BaseModel):
    enabled: bool


@app.put('/registration-settings')
def update_registration_settings(setting_data: RegistrationSettingUpdate, admin_user: Dict[str, Any] = Depends(require_permission('system_settings'))):
    """更新注册开关设置"""
    from db_manager import db_manager
    try:
        enabled = setting_data.enabled
        success = db_manager.set_system_setting(
            'registration_enabled',
            'true' if enabled else 'false',
            '是否开启用户注册'
        )
        if success:
            log_with_user('info', f"更新注册设置: {'开启' if enabled else '关闭'}", admin_user)
            return {
                'success': True,
                'enabled': enabled,
                'message': f"注册功能已{'开启' if enabled else '关闭'}"
            }
        else:
            raise HTTPException(status_code=500, detail='更新注册设置失败')
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新注册设置失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))





@app.delete("/cookies/{cid}")
def remove_cookie(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        cookie_manager.manager.remove_cookie(cid)
        return {"msg": "removed"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class AutoConfirmUpdate(BaseModel):
    auto_confirm: bool


class RemarkUpdate(BaseModel):
    remark: str


class PauseDurationUpdate(BaseModel):
    pause_duration: int


@app.put("/cookies/{cid}/auto-confirm")
def update_auto_confirm(cid: str, update_data: AutoConfirmUpdate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号的自动确认发货设置"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 更新数据库中的auto_confirm设置
        success = db_manager.update_auto_confirm(cid, update_data.auto_confirm)
        if not success:
            raise HTTPException(status_code=500, detail="更新自动确认发货设置失败")

        # 通知CookieManager更新设置（如果账号正在运行）
        if hasattr(cookie_manager.manager, 'update_auto_confirm_setting'):
            cookie_manager.manager.update_auto_confirm_setting(cid, update_data.auto_confirm)

        return {
            "msg": "success",
            "auto_confirm": update_data.auto_confirm,
            "message": f"自动确认发货已{'开启' if update_data.auto_confirm else '关闭'}"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/cookies/{cid}/auto-confirm")
def get_auto_confirm(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取账号的自动确认发货设置"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 获取auto_confirm设置
        auto_confirm = db_manager.get_auto_confirm(cid)
        return {
            "auto_confirm": auto_confirm,
            "message": f"自动确认发货当前{'开启' if auto_confirm else '关闭'}"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/cookies/{cid}/remark")
def update_cookie_remark(cid: str, update_data: RemarkUpdate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号备注"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 更新备注
        success = db_manager.update_cookie_remark(cid, update_data.remark)
        if success:
            log_with_user('info', f"更新账号备注: {cid} -> {update_data.remark}", current_user)
            return {
                "message": "备注更新成功",
                "remark": update_data.remark
            }
        else:
            raise HTTPException(status_code=500, detail="备注更新失败")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/cookies/{cid}/remark")
def get_cookie_remark(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取账号备注"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 获取Cookie详细信息（包含备注）
        cookie_details = db_manager.get_cookie_details(cid)
        if cookie_details:
            return {
                "remark": cookie_details.get('remark', ''),
                "message": "获取备注成功"
            }
        else:
            raise HTTPException(status_code=404, detail="账号不存在")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/cookies/{cid}/pause-duration")
def update_cookie_pause_duration(cid: str, update_data: PauseDurationUpdate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号自动回复暂停时间"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 验证暂停时间范围（1-60分钟）
        if not (1 <= update_data.pause_duration <= 60):
            raise HTTPException(status_code=400, detail="暂停时间必须在1-60分钟之间")

        # 更新暂停时间
        success = db_manager.update_cookie_pause_duration(cid, update_data.pause_duration)
        if success:
            log_with_user('info', f"更新账号自动回复暂停时间: {cid} -> {update_data.pause_duration}分钟", current_user)
            return {
                "message": "暂停时间更新成功",
                "pause_duration": update_data.pause_duration
            }
        else:
            raise HTTPException(status_code=500, detail="暂停时间更新失败")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/cookies/{cid}/pause-duration")
def get_cookie_pause_duration(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取账号自动回复暂停时间"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 获取暂停时间
        pause_duration = db_manager.get_cookie_pause_duration(cid)
        return {
            "pause_duration": pause_duration,
            "message": "获取暂停时间成功"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))




class KeywordIn(BaseModel):
    keywords: Dict[str, str]  # key -> reply

class KeywordWithItemIdIn(BaseModel):
    keywords: List[Dict[str, Any]]  # [{"keyword": str, "reply": str, "item_id": str}]


@app.get("/keywords/{cid}")
def get_keywords(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    # 直接从数据库获取所有关键词（避免重复计算）
    item_keywords = db_manager.get_keywords_with_item_id(cid)

    # 转换为统一格式
    all_keywords = []
    for keyword, reply, item_id in item_keywords:
        all_keywords.append({
            "keyword": keyword,
            "reply": reply,
            "item_id": item_id,
            "type": "item" if item_id else "normal"
        })

    return all_keywords


@app.get("/keywords-with-item-id/{cid}")
def get_keywords_with_item_id(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取包含商品ID的关键词列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    # 获取包含类型信息的关键词
    keywords = db_manager.get_keywords_with_type(cid)

    # 转换为前端需要的格式
    result = []
    for keyword_data in keywords:
        result.append({
            "keyword": keyword_data['keyword'],
            "reply": keyword_data['reply'],
            "item_id": keyword_data['item_id'] or "",
            "type": keyword_data['type'],
            "image_url": keyword_data['image_url']
        })

    return result


@app.post("/keywords/{cid}")
def update_keywords(cid: str, body: KeywordIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        log_with_user('warning', f"尝试操作其他用户的Cookie关键字: {cid}", current_user)
        raise HTTPException(status_code=403, detail="无权限操作该Cookie")

    kw_list = [(k, v) for k, v in body.keywords.items()]
    log_with_user('info', f"更新Cookie关键字: {cid}, 数量: {len(kw_list)}", current_user)

    cookie_manager.manager.update_keywords(cid, kw_list)
    log_with_user('info', f"Cookie关键字更新成功: {cid}", current_user)
    return {"msg": "updated", "count": len(kw_list)}


@app.post("/keywords-with-item-id/{cid}")
def update_keywords_with_item_id(cid: str, body: KeywordWithItemIdIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新包含商品ID的关键词列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        log_with_user('warning', f"尝试操作其他用户的Cookie关键字: {cid}", current_user)
        raise HTTPException(status_code=403, detail="无权限操作该Cookie")

    # 验证数据格式
    keywords_to_save = []
    keyword_set = set()  # 用于检查当前提交的关键词中是否有重复

    for kw_data in body.keywords:
        keyword = kw_data.get('keyword', '').strip()
        reply = kw_data.get('reply', '').strip()
        item_id = kw_data.get('item_id', '').strip() or None

        if not keyword:
            raise HTTPException(status_code=400, detail="关键词不能为空")

        # 检查当前提交的关键词中是否有重复
        keyword_key = f"{keyword}|{item_id or ''}"
        if keyword_key in keyword_set:
            item_id_text = f"（商品ID: {item_id}）" if item_id else "（通用关键词）"
            raise HTTPException(status_code=400, detail=f"关键词 '{keyword}' {item_id_text} 在当前提交中重复")
        keyword_set.add(keyword_key)

        keywords_to_save.append((keyword, reply, item_id))

    # 保存关键词（只保存文本关键词，保留图片关键词）
    try:
        success = db_manager.save_text_keywords_only(cid, keywords_to_save)
        if not success:
            raise HTTPException(status_code=500, detail="保存关键词失败")
    except Exception as e:
        error_msg = str(e)
        if "UNIQUE constraint failed" in error_msg:
            # 解析具体的冲突信息
            if "keywords.cookie_id, keywords.keyword" in error_msg:
                raise HTTPException(status_code=400, detail="关键词重复！该关键词已存在（可能是图片关键词或文本关键词），请使用其他关键词")
            else:
                raise HTTPException(status_code=400, detail="关键词重复！请使用不同的关键词或商品ID组合")
        else:
            log_with_user('error', f"保存关键词时发生未知错误: {error_msg}", current_user)
            raise HTTPException(status_code=500, detail="保存关键词失败")

    log_with_user('info', f"更新Cookie关键字(含商品ID): {cid}, 数量: {len(keywords_to_save)}", current_user)
    return {"msg": "updated", "count": len(keywords_to_save)}


@app.get("/items/{cid}")
def get_items_list(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定账号的商品列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    try:
        # 获取该账号的所有商品
        with db_manager.lock:
            cursor = db_manager.conn.cursor()
            cursor.execute('''
            SELECT item_id, item_title, item_price, updated_at
            FROM item_info
            WHERE cookie_id = ?
            ORDER BY updated_at DESC
            ''', (cid,))

            items = []
            for row in cursor.fetchall():
                items.append({
                    'item_id': row[0],
                    'item_title': row[1] or '未知商品',
                    'item_price': row[2] or '价格未知',
                    'updated_at': row[3]
                })

            return {"items": items, "count": len(items)}

    except Exception as e:
        logger.error(f"获取商品列表失败: {e}")
        raise HTTPException(status_code=500, detail="获取商品列表失败")


@app.get("/keywords-export/{cid}")
def export_keywords(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """导出指定账号的关键词为Excel文件"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    try:
        # 获取关键词数据（包含类型信息）
        keywords = db_manager.get_keywords_with_type(cid)

        # 创建DataFrame，只导出文本类型的关键词
        data = []
        for keyword_data in keywords:
            # 只导出文本类型的关键词
            if keyword_data.get('type', 'text') == 'text':
                data.append({
                    '关键词': keyword_data['keyword'],
                    '商品ID': keyword_data['item_id'] or '',
                    '关键词内容': keyword_data['reply']
                })

        # 如果没有数据，创建空的DataFrame但保留列名（作为模板）
        if not data:
            df = pd.DataFrame(columns=['关键词', '商品ID', '关键词内容'])
        else:
            df = pd.DataFrame(data)

        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='关键词数据', index=False)

            # 如果是空模板，添加一些示例说明
            if data == []:
                worksheet = writer.sheets['关键词数据']
                # 添加示例数据作为注释（从第2行开始）
                worksheet['A2'] = '你好'
                worksheet['B2'] = ''
                worksheet['C2'] = '您好！欢迎咨询，有什么可以帮助您的吗？'

                worksheet['A3'] = '价格'
                worksheet['B3'] = '123456'
                worksheet['C3'] = '这个商品的价格是99元，现在有优惠活动哦！'

                worksheet['A4'] = '发货'
                worksheet['B4'] = ''
                worksheet['C4'] = '我们会在24小时内发货，请耐心等待。'

                # 设置示例行的样式（浅灰色背景）
                from openpyxl.styles import PatternFill
                gray_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                for row in range(2, 5):
                    for col in range(1, 4):
                        worksheet.cell(row=row, column=col).fill = gray_fill

        output.seek(0)

        # 生成文件名（使用URL编码处理中文）
        from urllib.parse import quote
        if not data:
            filename = f"keywords_template_{cid}_{int(time.time())}.xlsx"
        else:
            filename = f"keywords_{cid}_{int(time.time())}.xlsx"
        encoded_filename = quote(filename.encode('utf-8'))

        # 返回文件
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )

    except Exception as e:
        logger.error(f"导出关键词失败: {e}")
        raise HTTPException(status_code=500, detail=f"导出关键词失败: {str(e)}")


@app.post("/keywords-import/{cid}")
async def import_keywords(cid: str, file: UploadFile = File(...), current_user: Dict[str, Any] = Depends(get_current_user)):
    """导入Excel文件中的关键词到指定账号"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    # 检查文件类型
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="请上传Excel文件(.xlsx或.xls)")

    try:
        # 读取Excel文件
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))

        # 检查必要的列
        required_columns = ['关键词', '商品ID', '关键词内容']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(status_code=400, detail=f"Excel文件缺少必要的列: {', '.join(missing_columns)}")

        # 获取现有的文本类型关键词（用于比较更新/新增）
        existing_keywords = db_manager.get_keywords_with_type(cid)
        existing_dict = {}
        for keyword_data in existing_keywords:
            # 只考虑文本类型的关键词
            if keyword_data.get('type', 'text') == 'text':
                keyword = keyword_data['keyword']
                reply = keyword_data['reply']
                item_id = keyword_data['item_id']
                key = f"{keyword}|{item_id or ''}"
                existing_dict[key] = (keyword, reply, item_id)

        # 处理导入数据
        import_data = []
        update_count = 0
        add_count = 0

        for index, row in df.iterrows():
            keyword = str(row['关键词']).strip()
            item_id = str(row['商品ID']).strip() if pd.notna(row['商品ID']) and str(row['商品ID']).strip() else None
            reply = str(row['关键词内容']).strip()

            if not keyword:
                continue  # 跳过没有关键词的行

            # 检查是否重复
            key = f"{keyword}|{item_id or ''}"
            if key in existing_dict:
                # 更新现有关键词
                update_count += 1
            else:
                # 新增关键词
                add_count += 1

            import_data.append((keyword, reply, item_id))

        if not import_data:
            raise HTTPException(status_code=400, detail="Excel文件中没有有效的关键词数据")

        # 保存到数据库（只影响文本关键词，保留图片关键词）
        success = db_manager.save_text_keywords_only(cid, import_data)
        if not success:
            raise HTTPException(status_code=500, detail="保存关键词到数据库失败")

        log_with_user('info', f"导入关键词成功: {cid}, 新增: {add_count}, 更新: {update_count}", current_user)

        return {
            "msg": "导入成功",
            "total": len(import_data),
            "added": add_count,
            "updated": update_count
        }

    except pd.errors.EmptyDataError:
        raise HTTPException(status_code=400, detail="Excel文件为空")
    except pd.errors.ParserError:
        raise HTTPException(status_code=400, detail="Excel文件格式错误")
    except Exception as e:
        logger.error(f"导入关键词失败: {e}")
        raise HTTPException(status_code=500, detail=f"导入关键词失败: {str(e)}")


@app.post("/keywords/{cid}/image")
async def add_image_keyword(
    cid: str,
    keyword: str = Form(...),
    item_id: str = Form(default=""),
    image: UploadFile = File(...),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """添加图片关键词"""
    logger.info(f"接收到图片关键词添加请求: cid={cid}, keyword={keyword}, item_id={item_id}")

    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查参数
    if not keyword or not keyword.strip():
        raise HTTPException(status_code=400, detail="关键词不能为空")

    if not image or not image.filename:
        raise HTTPException(status_code=400, detail="请选择图片文件")

    # 检查cookie是否属于当前用户
    cookie_details = db_manager.get_cookie_details(cid)
    if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
        raise HTTPException(status_code=404, detail="账号不存在或无权限")

    try:
        logger.info(f"接收到图片关键词添加请求: cid={cid}, keyword={keyword}, item_id={item_id}, filename={image.filename}")

        # 验证图片文件
        if not image.content_type or not image.content_type.startswith('image/'):
            logger.warning(f"无效的图片文件类型: {image.content_type}")
            raise HTTPException(status_code=400, detail="请上传图片文件")

        # 读取图片数据
        image_data = await image.read()
        logger.info(f"读取图片数据成功，大小: {len(image_data)} bytes")

        # 保存图片
        image_url = image_manager.save_image(image_data, image.filename)
        if not image_url:
            logger.error("图片保存失败")
            raise HTTPException(status_code=400, detail="图片保存失败")

        logger.info(f"图片保存成功: {image_url}")

        # 先检查关键词是否已存在
        normalized_item_id = item_id if item_id and item_id.strip() else None
        if db_manager.check_keyword_duplicate(cid, keyword, normalized_item_id):
            # 删除已保存的图片
            image_manager.delete_image(image_url)
            if normalized_item_id:
                raise HTTPException(status_code=400, detail=f"关键词 '{keyword}' 在商品 '{normalized_item_id}' 中已存在")
            else:
                raise HTTPException(status_code=400, detail=f"通用关键词 '{keyword}' 已存在")

        # 保存图片关键词到数据库
        success = db_manager.save_image_keyword(cid, keyword, image_url, item_id or None)
        if not success:
            # 如果数据库保存失败，删除已保存的图片
            logger.error("数据库保存失败，删除已保存的图片")
            image_manager.delete_image(image_url)
            raise HTTPException(status_code=400, detail="图片关键词保存失败，请稍后重试")

        log_with_user('info', f"添加图片关键词成功: {cid}, 关键词: {keyword}", current_user)

        return {
            "msg": "图片关键词添加成功",
            "keyword": keyword,
            "image_url": image_url,
            "item_id": item_id or None
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"添加图片关键词失败: {e}")
        raise HTTPException(status_code=500, detail=f"添加图片关键词失败: {str(e)}")


@app.post("/upload-image")
async def upload_image(
    image: UploadFile = File(...),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """上传图片（用于卡券等功能）"""
    try:
        logger.info(f"接收到图片上传请求: filename={image.filename}")

        # 验证图片文件
        if not image.content_type or not image.content_type.startswith('image/'):
            logger.warning(f"无效的图片文件类型: {image.content_type}")
            raise HTTPException(status_code=400, detail="请上传图片文件")

        # 读取图片数据
        image_data = await image.read()
        logger.info(f"读取图片数据成功，大小: {len(image_data)} bytes")

        # 保存图片
        image_url = image_manager.save_image(image_data, image.filename)
        if not image_url:
            logger.error("图片保存失败")
            raise HTTPException(status_code=400, detail="图片保存失败")

        logger.info(f"图片上传成功: {image_url}")

        return {
            "message": "图片上传成功",
            "image_url": image_url
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"图片上传失败: {e}")
        raise HTTPException(status_code=500, detail=f"图片上传失败: {str(e)}")


@app.get("/keywords-with-type/{cid}")
def get_keywords_with_type(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取包含类型信息的关键词列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    cookie_details = db_manager.get_cookie_details(cid)
    if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
        raise HTTPException(status_code=404, detail="账号不存在或无权限")

    try:
        keywords = db_manager.get_keywords_with_type(cid)
        return keywords
    except Exception as e:
        logger.error(f"获取关键词列表失败: {e}")
        raise HTTPException(status_code=500, detail=f"获取关键词列表失败: {str(e)}")


@app.delete("/keywords/{cid}/{index}")
def delete_keyword_by_index(cid: str, index: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """根据索引删除关键词"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    cookie_details = db_manager.get_cookie_details(cid)
    if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
        raise HTTPException(status_code=404, detail="账号不存在或无权限")

    try:
        # 先获取要删除的关键词信息（用于删除图片文件）
        keywords = db_manager.get_keywords_with_type(cid)
        if 0 <= index < len(keywords):
            keyword_data = keywords[index]

            # 删除关键词
            success = db_manager.delete_keyword_by_index(cid, index)
            if not success:
                raise HTTPException(status_code=400, detail="删除关键词失败")

            # 如果是图片关键词，删除对应的图片文件
            if keyword_data.get('type') == 'image' and keyword_data.get('image_url'):
                image_manager.delete_image(keyword_data['image_url'])

            log_with_user('info', f"删除关键词成功: {cid}, 索引: {index}, 关键词: {keyword_data.get('keyword')}", current_user)

            return {"msg": "删除成功"}
        else:
            raise HTTPException(status_code=400, detail="关键词索引无效")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除关键词失败: {e}")
        raise HTTPException(status_code=500, detail=f"删除关键词失败: {str(e)}")


@app.get("/debug/keywords-table-info")
def debug_keywords_table_info(current_user: Dict[str, Any] = Depends(get_current_user)):
    """调试：检查keywords表结构"""
    try:
        import sqlite3
        conn = sqlite3.connect(db_manager.db_path)
        cursor = conn.cursor()

        # 获取表结构信息
        cursor.execute("PRAGMA table_info(keywords)")
        columns = cursor.fetchall()

        # 获取数据库版本
        cursor.execute("SELECT value FROM system_settings WHERE key = 'db_version'")
        version_result = cursor.fetchone()
        db_version = version_result[0] if version_result else "未知"

        conn.close()

        return {
            "db_version": db_version,
            "table_columns": [{"name": col[1], "type": col[2], "default": col[4]} for col in columns]
        }
    except Exception as e:
        logger.error(f"检查表结构失败: {e}")
        raise HTTPException(status_code=500, detail=f"检查表结构失败: {str(e)}")


# 卡券管理API
@app.get("/cards")
def get_cards(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的卡券列表"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        cards = db_manager.get_all_cards(user_id)
        return cards
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/cards")
def create_card(card_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """创建新卡券"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        card_name = card_data.get('name', '未命名卡券')

        log_with_user('info', f"创建卡券: {card_name}", current_user)

        # 验证多规格字段
        is_multi_spec = card_data.get('is_multi_spec', False)
        if is_multi_spec:
            if not card_data.get('spec_name') or not card_data.get('spec_value'):
                raise HTTPException(status_code=400, detail="多规格卡券必须提供规格名称和规格值")

        card_id = db_manager.create_card(
            name=card_data.get('name'),
            card_type=card_data.get('type'),
            api_config=card_data.get('api_config'),
            text_content=card_data.get('text_content'),
            data_content=card_data.get('data_content'),
            image_url=card_data.get('image_url'),
            description=card_data.get('description'),
            enabled=card_data.get('enabled', True),
            delay_seconds=card_data.get('delay_seconds', 0),
            is_multi_spec=is_multi_spec,
            spec_name=card_data.get('spec_name') if is_multi_spec else None,
            spec_value=card_data.get('spec_value') if is_multi_spec else None,
            user_id=user_id
        )

        log_with_user('info', f"卡券创建成功: {card_name} (ID: {card_id})", current_user)
        return {"id": card_id, "message": "卡券创建成功"}
    except Exception as e:
        log_with_user('error', f"创建卡券失败: {card_data.get('name', '未知')} - {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/cards/{card_id}")
def get_card(card_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取单个卡券详情"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        card = db_manager.get_card_by_id(card_id, user_id)
        if card:
            return card
        else:
            raise HTTPException(status_code=404, detail="卡券不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/cards/{card_id}")
def update_card(card_id: int, card_data: dict, _: None = Depends(require_auth)):
    """更新卡券"""
    try:
        from db_manager import db_manager
        # 验证多规格字段
        is_multi_spec = card_data.get('is_multi_spec')
        if is_multi_spec:
            if not card_data.get('spec_name') or not card_data.get('spec_value'):
                raise HTTPException(status_code=400, detail="多规格卡券必须提供规格名称和规格值")

        success = db_manager.update_card(
            card_id=card_id,
            name=card_data.get('name'),
            card_type=card_data.get('type'),
            api_config=card_data.get('api_config'),
            text_content=card_data.get('text_content'),
            data_content=card_data.get('data_content'),
            image_url=card_data.get('image_url'),
            description=card_data.get('description'),
            enabled=card_data.get('enabled', True),
            delay_seconds=card_data.get('delay_seconds'),
            is_multi_spec=is_multi_spec,
            spec_name=card_data.get('spec_name'),
            spec_value=card_data.get('spec_value')
        )
        if success:
            return {"message": "卡券更新成功"}
        else:
            raise HTTPException(status_code=404, detail="卡券不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/cards/{card_id}/image")
async def update_card_with_image(
    card_id: int,
    image: UploadFile = File(...),
    name: str = Form(...),
    type: str = Form(...),
    description: str = Form(default=""),
    delay_seconds: int = Form(default=0),
    enabled: bool = Form(default=True),
    is_multi_spec: bool = Form(default=False),
    spec_name: str = Form(default=""),
    spec_value: str = Form(default=""),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """更新带图片的卡券"""
    try:
        logger.info(f"接收到带图片的卡券更新请求: card_id={card_id}, name={name}, type={type}")

        # 验证图片文件
        if not image.content_type or not image.content_type.startswith('image/'):
            logger.warning(f"无效的图片文件类型: {image.content_type}")
            raise HTTPException(status_code=400, detail="请上传图片文件")

        # 验证多规格字段
        if is_multi_spec:
            if not spec_name or not spec_value:
                raise HTTPException(status_code=400, detail="多规格卡券必须提供规格名称和规格值")

        # 读取图片数据
        image_data = await image.read()
        logger.info(f"读取图片数据成功，大小: {len(image_data)} bytes")

        # 保存图片
        image_url = image_manager.save_image(image_data, image.filename)
        if not image_url:
            logger.error("图片保存失败")
            raise HTTPException(status_code=400, detail="图片保存失败")

        logger.info(f"图片保存成功: {image_url}")

        # 更新卡券
        from db_manager import db_manager
        success = db_manager.update_card(
            card_id=card_id,
            name=name,
            card_type=type,
            image_url=image_url,
            description=description,
            enabled=enabled,
            delay_seconds=delay_seconds,
            is_multi_spec=is_multi_spec,
            spec_name=spec_name if is_multi_spec else None,
            spec_value=spec_value if is_multi_spec else None
        )

        if success:
            logger.info(f"卡券更新成功: {name} (ID: {card_id})")
            return {"message": "卡券更新成功", "image_url": image_url}
        else:
            # 如果数据库更新失败，删除已保存的图片
            image_manager.delete_image(image_url)
            raise HTTPException(status_code=404, detail="卡券不存在")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新带图片的卡券失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# 自动发货规则API
@app.get("/delivery-rules")
def get_delivery_rules(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取发货规则列表"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        rules = db_manager.get_all_delivery_rules(user_id)
        return rules
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/delivery-rules")
def create_delivery_rule(rule_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """创建新发货规则"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        rule_id = db_manager.create_delivery_rule(
            keyword=rule_data.get('keyword'),
            card_id=rule_data.get('card_id'),
            delivery_count=rule_data.get('delivery_count', 1),
            enabled=rule_data.get('enabled', True),
            description=rule_data.get('description'),
            user_id=user_id
        )
        return {"id": rule_id, "message": "发货规则创建成功"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/delivery-rules/{rule_id}")
def get_delivery_rule(rule_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取单个发货规则详情"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        rule = db_manager.get_delivery_rule_by_id(rule_id, user_id)
        if rule:
            return rule
        else:
            raise HTTPException(status_code=404, detail="发货规则不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/delivery-rules/{rule_id}")
def update_delivery_rule(rule_id: int, rule_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新发货规则"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        success = db_manager.update_delivery_rule(
            rule_id=rule_id,
            keyword=rule_data.get('keyword'),
            card_id=rule_data.get('card_id'),
            delivery_count=rule_data.get('delivery_count', 1),
            enabled=rule_data.get('enabled', True),
            description=rule_data.get('description'),
            user_id=user_id
        )
        if success:
            return {"message": "发货规则更新成功"}
        else:
            raise HTTPException(status_code=404, detail="发货规则不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/cards/{card_id}")
def delete_card(card_id: int, _: None = Depends(require_auth)):
    """删除卡券"""
    try:
        from db_manager import db_manager
        success = db_manager.delete_card(card_id)
        if success:
            return {"message": "卡券删除成功"}
        else:
            raise HTTPException(status_code=404, detail="卡券不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/delivery-rules/{rule_id}")
def delete_delivery_rule(rule_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """删除发货规则"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        success = db_manager.delete_delivery_rule(rule_id, user_id)
        if success:
            return {"message": "发货规则删除成功"}
        else:
            raise HTTPException(status_code=404, detail="发货规则不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== 备份和恢复 API ====================

@app.get("/backup/export")
def export_backup(current_user: Dict[str, Any] = Depends(get_current_user)):
    """导出用户备份"""
    try:
        from db_manager import db_manager
        user_id = current_user['user_id']
        username = current_user['username']

        # 导出当前用户的数据
        backup_data = db_manager.export_backup(user_id)

        # 生成文件名
        import datetime
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"xianyu_backup_{username}_{timestamp}.json"

        # 返回JSON响应，设置下载头
        response = JSONResponse(content=backup_data)
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Content-Type"] = "application/json"

        return response
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出备份失败: {str(e)}")


@app.post("/backup/import")
def import_backup(file: UploadFile = File(...), current_user: Dict[str, Any] = Depends(get_current_user)):
    """导入用户备份"""
    try:
        # 验证文件类型
        if not file.filename.endswith('.json'):
            raise HTTPException(status_code=400, detail="只支持JSON格式的备份文件")

        # 读取文件内容
        content = file.file.read()
        backup_data = json.loads(content.decode('utf-8'))

        # 导入备份到当前用户
        from db_manager import db_manager
        user_id = current_user['user_id']
        success = db_manager.import_backup(backup_data, user_id)

        if success:
            # 备份导入成功后，刷新 CookieManager 的内存缓存
            import cookie_manager
            if cookie_manager.manager:
                try:
                    cookie_manager.manager.reload_from_db()
                    logger.info("备份导入后已刷新 CookieManager 缓存")
                except Exception as e:
                    logger.error(f"刷新 CookieManager 缓存失败: {e}")

            return {"message": "备份导入成功"}
        else:
            raise HTTPException(status_code=400, detail="备份导入失败")

    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="备份文件格式无效")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导入备份失败: {str(e)}")


@app.post("/system/reload-cache")
def reload_cache(_: None = Depends(require_auth)):
    """重新加载系统缓存（用于手动刷新数据）"""
    try:
        import cookie_manager
        if cookie_manager.manager:
            success = cookie_manager.manager.reload_from_db()
            if success:
                return {"message": "系统缓存已刷新", "success": True}
            else:
                raise HTTPException(status_code=500, detail="缓存刷新失败")
        else:
            raise HTTPException(status_code=500, detail="CookieManager 未初始化")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"刷新缓存失败: {str(e)}")


# ==================== 商品管理 API ====================

@app.get("/items")
def get_all_items(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的所有商品信息"""
    try:
        # 只返回当前用户的商品信息
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        all_items = []
        for cookie_id in user_cookies.keys():
            items = db_manager.get_items_by_cookie(cookie_id)
            all_items.extend(items)

        # 按更新时间倒序排列
        all_items.sort(key=lambda x: x.get("updated_at") or "", reverse=True)

        return {"items": all_items}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品信息失败: {str(e)}")


# ==================== 商品搜索 API ====================

class ItemSearchRequest(BaseModel):
    keyword: str
    page: int = 1
    page_size: int = 20

class ItemSearchMultipleRequest(BaseModel):
    keyword: str
    total_pages: int = 1

@app.post("/items/search")
async def search_items(
    search_request: ItemSearchRequest,
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """搜索闲鱼商品"""
    user_info = f"【{current_user.get('username', 'unknown')}#{current_user.get('user_id', 'unknown')}】" if current_user else "【未登录】"

    try:
        logger.info(f"{user_info} 开始单页搜索: 关键词='{search_request.keyword}', 页码={search_request.page}, 每页={search_request.page_size}")

        from utils.item_search import search_xianyu_items

        # 执行搜索
        result = await search_xianyu_items(
            keyword=search_request.keyword,
            page=search_request.page,
            page_size=search_request.page_size
        )

        # 检查是否有错误
        has_error = result.get("error")
        items_count = len(result.get("items", []))

        logger.info(f"{user_info} 单页搜索完成: 获取到 {items_count} 条数据" +
                   (f", 错误: {has_error}" if has_error else ""))

        response_data = {
            "success": True,
            "data": result.get("items", []),
            "total": result.get("total", 0),
            "page": search_request.page,
            "page_size": search_request.page_size,
            "keyword": search_request.keyword,
            "is_real_data": result.get("is_real_data", False),
            "source": result.get("source", "unknown")
        }

        # 如果有错误信息，也包含在响应中
        if has_error:
            response_data["error"] = has_error

        return response_data

    except Exception as e:
        error_msg = str(e)
        logger.error(f"{user_info} 商品搜索失败: {error_msg}")
        raise HTTPException(status_code=500, detail=f"商品搜索失败: {error_msg}")


@app.get("/cookies/check")
async def check_valid_cookies(
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """检查是否有有效的cookies账户（必须是启用状态）"""
    try:
        if cookie_manager.manager is None:
            return {
                "success": True,
                "hasValidCookies": False,
                "validCount": 0,
                "enabledCount": 0,
                "totalCount": 0
            }

        from db_manager import db_manager

        # 获取所有cookies
        all_cookies = db_manager.get_all_cookies()

        # 检查启用状态和有效性
        valid_cookies = []
        enabled_cookies = []

        for cookie_id, cookie_value in all_cookies.items():
            # 检查是否启用
            is_enabled = cookie_manager.manager.get_cookie_status(cookie_id)
            if is_enabled:
                enabled_cookies.append(cookie_id)
                # 检查是否有效（长度大于50）
                if len(cookie_value) > 50:
                    valid_cookies.append(cookie_id)

        return {
            "success": True,
            "hasValidCookies": len(valid_cookies) > 0,
            "validCount": len(valid_cookies),
            "enabledCount": len(enabled_cookies),
            "totalCount": len(all_cookies)
        }

    except Exception as e:
        logger.error(f"检查cookies失败: {str(e)}")
        return {
            "success": False,
            "hasValidCookies": False,
            "error": str(e)
        }

@app.get("/cookies/online-status")
async def get_cookies_online_status(
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """获取所有Cookie的WebSocket实时在线状态"""
    try:
        if cookie_manager.manager is None:
            return {"success": True, "status": {}}

        from db_manager import db_manager as db
        user_id = current_user['user_id'] if current_user else None
        all_cookies = db.get_all_cookies(user_id)

        result = {}
        for cookie_id in all_cookies:
            is_enabled = cookie_manager.manager.get_cookie_status(cookie_id)
            is_online = cookie_manager.manager.get_online_status(cookie_id)
            # 禁用状态显示为 disabled，启用状态根据在线状态显示
            if not is_enabled:
                status = "disabled"
            elif is_online is True:
                status = "online"
            else:
                status = "offline"
            result[cookie_id] = {
                "status": status,
                "enabled": is_enabled,
                "online": is_online if is_online is not None else False
            }

        return {"success": True, "status": result}
    except Exception as e:
        logger.error(f"获取在线状态失败: {str(e)}")
        return {"success": False, "status": {}}

@app.post("/items/search_multiple")
async def search_multiple_pages(
    search_request: ItemSearchMultipleRequest,
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """搜索多页闲鱼商品"""
    user_info = f"【{current_user.get('username', 'unknown')}#{current_user.get('user_id', 'unknown')}】" if current_user else "【未登录】"

    try:
        logger.info(f"{user_info} 开始多页搜索: 关键词='{search_request.keyword}', 页数={search_request.total_pages}")

        from utils.item_search import search_multiple_pages_xianyu

        # 执行多页搜索
        result = await search_multiple_pages_xianyu(
            keyword=search_request.keyword,
            total_pages=search_request.total_pages
        )

        # 检查是否有错误
        has_error = result.get("error")
        items_count = len(result.get("items", []))

        logger.info(f"{user_info} 多页搜索完成: 获取到 {items_count} 条数据" +
                   (f", 错误: {has_error}" if has_error else ""))

        response_data = {
            "success": True,
            "data": result.get("items", []),
            "total": result.get("total", 0),
            "total_pages": search_request.total_pages,
            "keyword": search_request.keyword,
            "is_real_data": result.get("is_real_data", False),
            "is_fallback": result.get("is_fallback", False),
            "source": result.get("source", "unknown")
        }

        # 如果有错误信息，也包含在响应中
        if has_error:
            response_data["error"] = has_error

        return response_data

    except Exception as e:
        error_msg = str(e)
        logger.error(f"{user_info} 多页商品搜索失败: {error_msg}")
        raise HTTPException(status_code=500, detail=f"多页商品搜索失败: {error_msg}")


@app.get("/items/detail/{item_id}")
async def get_public_item_detail(
    item_id: str,
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """获取公开商品详情（通过外部API）"""
    try:
        from utils.item_search import get_item_detail_from_api

        # 从外部API获取商品详情
        detail = await get_item_detail_from_api(item_id)

        if detail:
            return {
                "success": True,
                "data": detail
            }
        else:
            raise HTTPException(status_code=404, detail="商品详情获取失败")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取商品详情失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"获取商品详情失败: {str(e)}")


@app.get("/items/cookie/{cookie_id}")
def get_items_by_cookie(cookie_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定Cookie的商品信息"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        items = db_manager.get_items_by_cookie(cookie_id)
        return {"items": items}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品信息失败: {str(e)}")


@app.get("/items/{cookie_id}/{item_id}")
def get_item_detail(cookie_id: str, item_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取商品详情"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        item = db_manager.get_item_info(cookie_id, item_id)
        if not item:
            raise HTTPException(status_code=404, detail="商品不存在")
        return {"item": item}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品详情失败: {str(e)}")


class ItemDetailUpdate(BaseModel):
    item_title: Optional[str] = None
    item_price: Optional[str] = None
    item_description: Optional[str] = None
    item_detail: Optional[str] = None
    sync_to_xianyu: bool = False


@app.put("/items/{cookie_id}/{item_id}")
def update_item_detail(
    cookie_id: str,
    item_id: str,
    update_data: ItemDetailUpdate,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """更新商品详情，可选择同步到闲鱼"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 保存到本地数据库
        success = db_manager.update_item_fields(
            cookie_id, item_id,
            item_title=update_data.item_title,
            item_price=update_data.item_price,
            item_description=update_data.item_description,
            item_detail=update_data.item_detail
        )
        if not success:
            raise HTTPException(status_code=400, detail="更新失败，商品不存在")

        result = {"message": "商品更新成功", "sync_success": True, "sync_result": ""}

        # 如果需要同步到闲鱼
        if update_data.sync_to_xianyu:
            has_sync_fields = any([
                update_data.item_title,
                update_data.item_price,
                update_data.item_description
            ])
            if has_sync_fields:
                # 闲鱼网页版不支持商品编辑，只能通过 APP 操作
                logger.info(f"闲鱼同步请求: {cookie_id}/{item_id} title={update_data.item_title}")
                result['sync_result'] = '闲鱼网页版不支持编辑商品，请到闲鱼 APP 修改'
                result['sync_success'] = False
            else:
                result['sync_result'] = '未修改标题/价格/描述，跳过同步'
        else:
            result['sync_result'] = '未勾选同步到闲鱼'

        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"更新商品详情失败: {str(e)}")


@app.delete("/items/{cookie_id}/{item_id}")
def delete_item_info(
    cookie_id: str,
    item_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """删除商品信息"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        success = db_manager.delete_item_info(cookie_id, item_id)
        if success:
            return {"message": "商品信息删除成功"}
        else:
            raise HTTPException(status_code=404, detail="商品信息不存在")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除商品信息异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


class BatchDeleteRequest(BaseModel):
    items: List[dict]  # [{"cookie_id": "xxx", "item_id": "yyy"}, ...]


class AIReplySettings(BaseModel):
    ai_enabled: bool
    model_name: str = "qwen-plus"
    api_key: str = ""
    base_url: str = "https://dashscope.aliyuncs.com/compatible-mode/v1"
    max_discount_percent: int = 10
    max_discount_amount: int = 100
    max_bargain_rounds: int = 3
    custom_prompts: str = ""
    refund_policy: str = "allow"


@app.delete("/items/batch")
def batch_delete_items(
    request: BatchDeleteRequest,
    _: None = Depends(require_auth)
):
    """批量删除商品信息"""
    try:
        if not request.items:
            raise HTTPException(status_code=400, detail="删除列表不能为空")

        success_count = db_manager.batch_delete_item_info(request.items)
        total_count = len(request.items)

        return {
            "message": f"批量删除完成",
            "success_count": success_count,
            "total_count": total_count,
            "failed_count": total_count - success_count
        }
    except Exception as e:
        logger.error(f"批量删除商品信息异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


# ==================== AI回复管理API ====================

@app.get("/ai-reply-settings/{cookie_id}")
def get_ai_reply_settings(cookie_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定账号的AI回复设置"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        settings = db_manager.get_ai_reply_settings(cookie_id)

        # 掩码处理 API 密钥：仅显示首4位+****+末4位
        api_key = settings.get('api_key', '')
        if api_key and len(api_key) > 8:
            masked_key = api_key[:4] + '****' + api_key[-4:]
        elif api_key:
            masked_key = api_key[:2] + '****' if len(api_key) > 4 else '****'
        else:
            masked_key = ''
        settings['api_key'] = masked_key

        return settings
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取AI回复设置异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


@app.put("/ai-reply-settings/{cookie_id}")
def update_ai_reply_settings(cookie_id: str, settings: AIReplySettings, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新指定账号的AI回复设置"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 检查账号是否存在
        if cookie_manager.manager is None:
            raise HTTPException(status_code=500, detail='CookieManager 未就绪')

        # 保存设置
        settings_dict = settings.dict()

        # 如果 API 密钥包含 ****，说明用户没改密钥，保持原值
        if settings_dict.get('api_key') and '****' in settings_dict['api_key']:
            # 从数据库获取当前密钥
            current_settings = db_manager.get_ai_reply_settings(cookie_id)
            if current_settings and current_settings.get('api_key'):
                settings_dict['api_key'] = current_settings['api_key']

        success = db_manager.save_ai_reply_settings(cookie_id, settings_dict)

        if success:
            # 清理客户端缓存，强制重新创建
            ai_reply_engine.clear_client_cache(cookie_id)

            # 如果启用了AI回复，记录日志
            if settings.ai_enabled:
                logger.info(f"账号 {cookie_id} 启用AI回复")
            else:
                logger.info(f"账号 {cookie_id} 禁用AI回复")

            return {"message": "AI回复设置更新成功"}
        else:
            raise HTTPException(status_code=400, detail="更新失败")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新AI回复设置异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


@app.get("/ai-reply-settings")
def get_all_ai_reply_settings(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户所有账号的AI回复设置"""
    try:
        # 只返回当前用户的AI回复设置
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        all_settings = db_manager.get_all_ai_reply_settings()
        # 过滤只属于当前用户的设置
        user_settings = {cid: settings for cid, settings in all_settings.items() if cid in user_cookies}
        return user_settings
    except Exception as e:
        logger.error(f"获取所有AI回复设置异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


@app.post("/ai-reply-test/{cookie_id}")
def test_ai_reply(cookie_id: str, test_data: dict = Body(...), _: None = Depends(require_auth)):
    """测试AI回复功能"""
    try:
        # 检查账号是否存在
        if cookie_manager.manager is None:
            raise HTTPException(status_code=500, detail='CookieManager 未就绪')

        if cookie_id not in cookie_manager.manager.cookies:
            raise HTTPException(status_code=404, detail='账号不存在')

        # 检查是否启用AI回复
        if not ai_reply_engine.is_ai_enabled(cookie_id):
            raise HTTPException(status_code=400, detail='该账号未启用AI回复')

        # 检查 API Key 是否已配置
        settings = db_manager.get_ai_reply_settings(cookie_id)
        if not settings.get('api_key'):
            raise HTTPException(status_code=400, detail='请先配置 AI 回复的 API Key')

        # 构造测试数据
        test_message = test_data.get('message', '你好')
        test_item_info = {
            'title': test_data.get('item_title', '测试商品'),
            'price': test_data.get('item_price', 100),
            'desc': test_data.get('item_desc', '这是一个测试商品')
        }

        # 生成测试回复
        reply = ai_reply_engine.generate_reply(
            message=test_message,
            item_info=test_item_info,
            chat_id=f"test_{int(time.time())}",
            cookie_id=cookie_id,
            user_id="test_user",
            item_id="test_item"
        )

        if reply:
            return {"message": "测试成功", "reply": reply}
        else:
            raise HTTPException(status_code=400, detail="AI回复生成失败")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"测试AI回复异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


# ------------------------- AI对话历史接口 -------------------------

@app.get("/ai-conversations")
def get_ai_conversations(
    cookie_id: str = None,
    buyer_id: str = None,
    item_id: str = None,
    page: int = 1,
    page_size: int = 20,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """分页获取AI对话历史"""
    try:
        from db_manager import db_manager
        uid = current_user['user_id']
        is_admin = current_user.get('is_admin', False)

        # 非管理员只能查看自己账号的对话
        if is_admin:
            user_cookie_ids = None
        else:
            user_cookies = db_manager.get_all_cookies(uid)
            user_cookie_ids = list(user_cookies.keys()) if user_cookies else []

        result = db_manager.get_ai_conversations(
            cookie_id=cookie_id,
            user_id=buyer_id,
            item_id=item_id,
            page=page,
            page_size=page_size,
            user_cookie_ids=user_cookie_ids
        )
        return result
    except Exception as e:
        logger.error(f"获取AI对话历史异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


# ==================== 日志管理API ====================

@app.get("/logs")
async def get_logs(lines: int = 200, level: str = None, source: str = None, _: None = Depends(require_auth)):
    """获取实时系统日志"""
    try:
        # 获取文件日志收集器
        collector = get_file_log_collector()

        # 获取日志
        logs = collector.get_logs(lines=lines, level_filter=level, source_filter=source)

        return {"success": True, "logs": logs}

    except Exception as e:
        return {"success": False, "message": f"获取日志失败: {str(e)}", "logs": []}


@app.get("/logs/stats")
async def get_log_stats(_: None = Depends(require_auth)):
    """获取日志统计信息"""
    try:
        collector = get_file_log_collector()
        stats = collector.get_stats()

        return {"success": True, "stats": stats}

    except Exception as e:
        return {"success": False, "message": f"获取日志统计失败: {str(e)}", "stats": {}}


@app.post("/logs/clear")
async def clear_logs(_: None = Depends(require_auth)):
    """清空日志"""
    try:
        collector = get_file_log_collector()
        collector.clear_logs()

        return {"success": True, "message": "日志已清空"}

    except Exception as e:
        return {"success": False, "message": f"清空日志失败: {str(e)}"}


# ==================== 商品管理API ====================

@app.post("/items/get-all-from-account")
async def get_all_items_from_account(request: dict, _: None = Depends(require_auth)):
    """从指定账号获取所有商品信息"""
    try:
        cookie_id = request.get('cookie_id')
        if not cookie_id:
            return {"success": False, "message": "缺少cookie_id参数"}

        # 获取指定账号的cookie信息
        cookie_info = db_manager.get_cookie_by_id(cookie_id)
        if not cookie_info:
            return {"success": False, "message": "未找到指定的账号信息"}

        cookies_str = cookie_info.get('cookies_str', '')
        if not cookies_str:
            return {"success": False, "message": "账号cookie信息为空"}

        # 创建XianyuLive实例，传入正确的cookie_id
        from XianyuAutoAsync import XianyuLive
        xianyu_instance = XianyuLive(cookies_str, cookie_id)

        # 调用获取所有商品信息的方法（自动分页）
        logger.info(f"开始获取账号 {cookie_id} 的所有商品信息")
        result = await xianyu_instance.get_all_items()

        # 关闭session
        await xianyu_instance.close_session()

        if result.get('error'):
            logger.error(f"获取商品信息失败: {result['error']}")
            return {"success": False, "message": result['error']}
        else:
            total_count = result.get('total_count', 0)
            total_pages = result.get('total_pages', 1)
            logger.info(f"成功获取账号 {cookie_id} 的 {total_count} 个商品（共{total_pages}页）")
            return {
                "success": True,
                "message": f"成功获取 {total_count} 个商品（共{total_pages}页），详细信息已打印到控制台",
                "total_count": total_count,
                "total_pages": total_pages
            }

    except Exception as e:
        logger.error(f"获取账号商品信息异常: {str(e)}")
        return {"success": False, "message": f"获取商品信息异常: {str(e)}"}


@app.post("/items/get-by-page")
async def get_items_by_page(request: dict, _: None = Depends(require_auth)):
    """从指定账号按页获取商品信息"""
    try:
        # 验证参数
        cookie_id = request.get('cookie_id')
        page_number = request.get('page_number', 1)
        page_size = request.get('page_size', 20)

        if not cookie_id:
            return {"success": False, "message": "缺少cookie_id参数"}

        # 验证分页参数
        try:
            page_number = int(page_number)
            page_size = int(page_size)
        except (ValueError, TypeError):
            return {"success": False, "message": "页码和每页数量必须是数字"}

        if page_number < 1:
            return {"success": False, "message": "页码必须大于0"}

        if page_size < 1 or page_size > 100:
            return {"success": False, "message": "每页数量必须在1-100之间"}

        # 获取账号信息
        account = db_manager.get_cookie_by_id(cookie_id)
        if not account:
            return {"success": False, "message": "账号不存在"}

        cookies_str = account['cookies_str']
        if not cookies_str:
            return {"success": False, "message": "账号cookies为空"}

        # 创建XianyuLive实例，传入正确的cookie_id
        from XianyuAutoAsync import XianyuLive
        xianyu_instance = XianyuLive(cookies_str, cookie_id)

        # 调用获取指定页商品信息的方法
        logger.info(f"开始获取账号 {cookie_id} 第{page_number}页商品信息（每页{page_size}条）")
        result = await xianyu_instance.get_item_list_info(page_number, page_size)

        # 关闭session
        await xianyu_instance.close_session()

        if result.get('error'):
            logger.error(f"获取商品信息失败: {result['error']}")
            return {"success": False, "message": result['error']}
        else:
            current_count = result.get('current_count', 0)
            logger.info(f"成功获取账号 {cookie_id} 第{page_number}页 {current_count} 个商品")
            return {
                "success": True,
                "message": f"成功获取第{page_number}页 {current_count} 个商品，详细信息已打印到控制台",
                "page_number": page_number,
                "page_size": page_size,
                "current_count": current_count
            }

    except Exception as e:
        logger.error(f"获取账号商品信息异常: {str(e)}")
        return {"success": False, "message": f"获取商品信息异常: {str(e)}"}


# ------------------------- 用户设置接口 -------------------------

@app.get('/user-settings')
def get_user_settings(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的设置"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        settings = db_manager.get_user_settings(user_id)
        return settings
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put('/user-settings/{key}')
def update_user_setting(key: str, setting_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新用户设置"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        value = setting_data.get('value')
        description = setting_data.get('description', '')

        log_with_user('info', f"更新用户设置: {key} = {value}", current_user)

        success = db_manager.set_user_setting(user_id, key, value, description)
        if success:
            log_with_user('info', f"用户设置更新成功: {key}", current_user)
            return {'msg': 'setting updated', 'key': key, 'value': value}
        else:
            log_with_user('error', f"用户设置更新失败: {key}", current_user)
            raise HTTPException(status_code=400, detail='更新失败')
    except Exception as e:
        log_with_user('error', f"更新用户设置异常: {key} - {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/user-settings/{key}')
def get_user_setting(key: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取用户特定设置"""
    from db_manager import db_manager
    try:
        user_id = current_user['user_id']
        setting = db_manager.get_user_setting(user_id, key)
        if setting:
            return setting
        else:
            raise HTTPException(status_code=404, detail='设置不存在')
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 管理员专用接口 -------------------------

@app.get('/admin/users')
def get_all_users(admin_user: Dict[str, Any] = Depends(require_permission('user_management'))):
    """获取所有用户信息（管理员专用）"""
    from db_manager import db_manager
    try:
        log_with_user('info', "查询所有用户信息", admin_user)
        users = db_manager.get_all_users()

        # 为每个用户添加统计信息
        for user in users:
            user_id = user['id']
            # 统计用户的Cookie数量
            user_cookies = db_manager.get_all_cookies(user_id)
            user['cookie_count'] = len(user_cookies)

            # 统计用户的卡券数量
            user_cards = db_manager.get_all_cards(user_id)
            user['card_count'] = len(user_cards) if user_cards else 0

            # 隐藏密码字段
            if 'password_hash' in user:
                del user['password_hash']

        log_with_user('info', f"返回用户信息，共 {len(users)} 个用户", admin_user)
        return {"users": users}
    except Exception as e:
        log_with_user('error', f"获取用户信息失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.post('/admin/users')
def create_user_by_admin(admin_user: Dict[str, Any] = Depends(require_permission('user_management')), data: dict = Body(...)):
    """管理员创建新用户"""
    from db_manager import db_manager
    try:
        username = data.get('username', '').strip()
        email = data.get('email', '').strip()
        password = data.get('password', '').strip()
        role = data.get('role', 'user').strip()

        if not username or not email or not password:
            raise HTTPException(status_code=400, detail="用户名、邮箱和密码不能为空")

        if len(username) < 2:
            raise HTTPException(status_code=400, detail="用户名至少2个字符")

        if len(password) < 6:
            raise HTTPException(status_code=400, detail="密码至少6位")

        if role not in ('user', 'admin'):
            raise HTTPException(status_code=400, detail="角色无效")

        # 检查用户名是否已存在
        existing = db_manager.get_user_by_username(username)
        if existing:
            raise HTTPException(status_code=400, detail=f"用户名 '{username}' 已存在")

        success = db_manager.create_user(username, email, password, role)
        if success:
            log_with_user('info', f"管理员创建用户成功: {username}, 角色: {role}", admin_user)
            return {"success": True, "message": f"用户 '{username}' 创建成功"}
        else:
            raise HTTPException(status_code=400, detail="创建用户失败，用户名或邮箱可能已存在")
    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"创建用户异常: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.put('/admin/users/{user_id}')
def update_user_by_admin(user_id: int, data: dict = Body(...), admin_user: Dict[str, Any] = Depends(require_permission('user_management'))):
    """管理员修改用户信息（密码重置）"""
    from db_manager import db_manager
    try:
        user = db_manager.get_user_by_id(user_id)
        if not user:
            raise HTTPException(status_code=404, detail="用户不存在")

        new_password = data.get('password', '').strip()
        if new_password:
            if len(new_password) < 6:
                raise HTTPException(status_code=400, detail="密码至少6位")
            success = db_manager.update_user_password(user['username'], new_password)
            if success:
                log_with_user('info', f"管理员重置用户密码成功: {user['username']}", admin_user)
                return {"success": True, "message": f"用户 '{user['username']}' 密码已重置"}

        new_email = data.get('email', '').strip()
        if new_email:
            db_manager.update_user_email(user['username'], new_email)
            log_with_user('info', f"管理员更新用户邮箱: {user['username']} -> {new_email}", admin_user)
            return {"success": True, "message": f"用户 '{user['username']}' 邮箱已更新"}

        new_role = data.get('role', '').strip()
        if new_role:
            if new_role not in ('user', 'admin'):
                raise HTTPException(status_code=400, detail="角色无效")
            db_manager.update_user_role(user['username'], new_role)
            log_with_user('info', f"管理员更新用户角色: {user['username']} -> {new_role}", admin_user)
            return {"success": True, "message": f"用户 '{user['username']}' 角色已更新为 {new_role}"}

        raise HTTPException(status_code=400, detail="未提供需要修改的字段")
    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"修改用户异常: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.delete('/admin/users/{user_id}')
def delete_user(user_id: int, admin_user: Dict[str, Any] = Depends(require_permission('user_management'))):
    """删除用户（管理员专用）"""
    from db_manager import db_manager
    try:
        # 不能删除管理员自己
        if user_id == admin_user['user_id']:
            log_with_user('warning', "尝试删除管理员自己", admin_user)
            raise HTTPException(status_code=400, detail="不能删除管理员自己")

        # 获取要删除的用户信息
        user_to_delete = db_manager.get_user_by_id(user_id)
        if not user_to_delete:
            raise HTTPException(status_code=404, detail="用户不存在")

        log_with_user('info', f"准备删除用户: {user_to_delete['username']} (ID: {user_id})", admin_user)

        # 删除用户及其相关数据
        success = db_manager.delete_user_and_data(user_id)

        if success:
            log_with_user('info', f"用户删除成功: {user_to_delete['username']} (ID: {user_id})", admin_user)
            return {"message": f"用户 {user_to_delete['username']} 删除成功"}
        else:
            log_with_user('error', f"用户删除失败: {user_to_delete['username']} (ID: {user_id})", admin_user)
            raise HTTPException(status_code=400, detail="删除失败")
    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"删除用户异常: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/admin/logs')
def get_system_logs(admin_user: Dict[str, Any] = Depends(require_permission('system_logs')),
                   lines: int = 100,
                   level: str = None):
    """获取系统日志（管理员专用）"""
    import os
    import glob
    from datetime import datetime

    try:
        log_with_user('info', f"查询系统日志，行数: {lines}, 级别: {level}", admin_user)

        # 查找日志文件
        log_files = glob.glob("logs/xianyu_*.log")
        logger.info(f"找到日志文件: {log_files}")

        if not log_files:
            logger.warning("未找到日志文件")
            return {"logs": [], "message": "未找到日志文件", "success": False}

        # 获取最新的日志文件
        latest_log_file = max(log_files, key=os.path.getctime)
        logger.info(f"使用最新日志文件: {latest_log_file}")

        logs = []
        try:
            with open(latest_log_file, 'r', encoding='utf-8') as f:
                all_lines = f.readlines()
                logger.info(f"读取到 {len(all_lines)} 行日志")

                # 如果指定了日志级别，进行过滤
                if level:
                    import re as re_module
                    level_pattern = re_module.compile(rf'\|\s*{re_module.escape(level.upper())}\s*\|')
                    filtered_lines = [line for line in all_lines if level_pattern.search(line)]
                    logger.info(f"按级别 {level} 过滤后剩余 {len(filtered_lines)} 行")
                else:
                    filtered_lines = all_lines

                # 获取最后N行
                recent_lines = filtered_lines[-lines:] if len(filtered_lines) > lines else filtered_lines
                logger.info(f"取最后 {len(recent_lines)} 行日志")

                for line in recent_lines:
                    logs.append(line.strip())

        except Exception as e:
            logger.error(f"读取日志文件失败: {str(e)}")
            log_with_user('error', f"读取日志文件失败: {str(e)}", admin_user)
            return {"logs": [], "message": f"读取日志文件失败: {str(e)}", "success": False}

        log_with_user('info', f"返回日志记录 {len(logs)} 条", admin_user)
        logger.info(f"成功返回 {len(logs)} 条日志记录")

        return {
            "logs": logs,
            "log_file": latest_log_file,
            "total_lines": len(logs),
            "success": True
        }

    except Exception as e:
        logger.error(f"获取系统日志失败: {str(e)}")
        log_with_user('error', f"获取系统日志失败: {str(e)}", admin_user)
        return {"logs": [], "message": f"获取系统日志失败: {str(e)}", "success": False}

@app.get('/admin/stats')
def get_system_stats(admin_user: Dict[str, Any] = Depends(require_permission('data_management'))):
    """获取系统统计信息"""
    from db_manager import db_manager
    try:
        log_with_user('info', "查询系统统计信息", admin_user)

        stats = {
            "users": {
                "total": 0,
                "active_today": 0
            },
            "cookies": {
                "total": 0,
                "enabled": 0
            },
            "cards": {
                "total": 0,
                "enabled": 0
            },
            "system": {
                "uptime": "未知",
                "version": "1.0.0"
            }
        }

        # 用户统计
        all_users = db_manager.get_all_users()
        stats["users"]["total"] = len(all_users)

        # Cookie统计
        all_cookies = db_manager.get_all_cookies()
        stats["cookies"]["total"] = len(all_cookies)

        # 卡券统计
        all_cards = db_manager.get_all_cards()
        if all_cards:
            stats["cards"]["total"] = len(all_cards)
            stats["cards"]["enabled"] = len([card for card in all_cards if card.get('enabled', True)])

        log_with_user('info', "系统统计信息查询完成", admin_user)
        return stats

    except Exception as e:
        log_with_user('error', f"获取系统统计信息失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/admin/permissions')
def get_default_roles() -> dict:
    """获取默认角色权限配置"""
    admin_perms = {k: True for k in ALL_PERMISSIONS}
    user_perms = {}
    for k in ALL_PERMISSIONS:
        # 管理员功能默认不给普通用户
        if k in ('user_management', 'system_logs', 'data_management', 'system_settings', 'export_backup', 'import_backup'):
            user_perms[k] = False
        else:
            user_perms[k] = True
    return {'admin': admin_perms, 'user': user_perms}

def get_permissions(admin_user: Dict[str, Any] = Depends(require_admin)):
    """获取角色权限配置"""
    from db_manager import db_manager as _db
    try:
        val = _db.get_system_setting('user_permissions')
        roles = json.loads(val) if val else {}
        if not roles:
            roles = get_default_roles()
            _db.set_system_setting('user_permissions', json.dumps(roles), '角色权限配置')
        return {"success": True, "roles": roles, "all_permissions": ALL_PERMISSIONS}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put('/admin/permissions')
def set_permissions(data: dict = Body(...), admin_user: Dict[str, Any] = Depends(require_admin)):
    """保存角色权限配置"""
    from db_manager import db_manager as _db
    try:
        roles = data.get('roles', {})
        _db.set_system_setting('user_permissions', json.dumps(roles), '角色权限配置')
        log_with_user('info', f"更新角色权限配置", admin_user)
        return {"success": True, "message": "角色权限配置已保存"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/admin/roles')
def get_roles(admin_user: Dict[str, Any] = Depends(require_admin)):
    """获取所有可用角色列表"""
    from db_manager import db_manager as _db
    try:
        val = _db.get_system_setting('user_permissions')
        roles = json.loads(val) if val else {}
        role_names = list(roles.keys())
        if 'admin' not in role_names:
            role_names.insert(0, 'admin')
        if 'user' not in role_names:
            role_names.append('user')
        return {"success": True, "roles": role_names}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/my-permissions')
def get_my_permissions(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的有效权限"""
    try:
        username = current_user.get('username', '')
        if username == 'admin':
            return {"success": True, "role": "admin", "permissions": {k: True for k in ALL_PERMISSIONS}}

        role = current_user.get('role', 'user')
        roles_config = get_user_permissions()
        role_data = roles_config.get(role, {})
        if role == 'admin' and not role_data:
            return {"success": True, "role": role, "permissions": {k: True for k in ALL_PERMISSIONS}}

        return {"success": True, "role": role, "permissions": role_data}
    except Exception as e:
        return {"success": False, "error": str(e), "permissions": {}}

# ------------------------- 指定商品回复接口 -------------------------

@app.get("/itemReplays")
def get_all_items(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的所有商品回复信息"""
    try:
        # 只返回当前用户的商品信息
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        all_items = []
        for cookie_id in user_cookies.keys():
            items = db_manager.get_itemReplays_by_cookie(cookie_id)
            all_items.extend(items)

        return {"items": all_items}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品回复信息失败: {str(e)}")

@app.get("/itemReplays/cookie/{cookie_id}")
def get_items_by_cookie(cookie_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定Cookie的商品信息"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        items = db_manager.get_itemReplays_by_cookie(cookie_id)
        return {"items": items}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品信息失败: {str(e)}")

@app.put("/item-reply/{cookie_id}/{item_id}")
def update_item_reply(
    cookie_id: str,
    item_id: str,
    data: dict,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    更新指定账号和商品的回复内容
    """
    try:
        user_id = current_user['user_id']
        from db_manager import db_manager

        # 验证cookie是否属于用户
        user_cookies = db_manager.get_all_cookies(user_id)
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        reply_content = data.get("reply_content", "").strip()
        if not reply_content:
            raise HTTPException(status_code=400, detail="回复内容不能为空")

        db_manager.update_item_reply(cookie_id=cookie_id, item_id=item_id, reply_content=reply_content)

        return {"message": "商品回复更新成功"}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"更新商品回复失败: {str(e)}")

@app.delete("/item-reply/{cookie_id}/{item_id}")
def delete_item_reply(cookie_id: str, item_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """
    删除指定账号cookie_id和商品item_id的商品回复
    """
    try:
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        success = db_manager.delete_item_reply(cookie_id, item_id)
        if not success:
            raise HTTPException(status_code=404, detail="商品回复不存在")

        return {"message": "商品回复删除成功"}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"删除商品回复失败: {str(e)}")

class ItemToDelete(BaseModel):
    cookie_id: str
    item_id: str

class BatchDeleteRequest(BaseModel):
    items: List[ItemToDelete]

@app.delete("/item-reply/batch")
async def batch_delete_item_reply(
    req: BatchDeleteRequest,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    批量删除商品回复
    """
    user_id = current_user['user_id']
    from db_manager import db_manager

    # 先校验当前用户是否有权限删除每个cookie对应的回复
    user_cookies = db_manager.get_all_cookies(user_id)
    for item in req.items:
        if item.cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail=f"无权限访问Cookie {item.cookie_id}")

    result = db_manager.batch_delete_item_replies([item.dict() for item in req.items])
    return {
        "success_count": result["success_count"],
        "failed_count": result["failed_count"]
    }

@app.get("/item-reply/{cookie_id}/{item_id}")
def get_item_reply(cookie_id: str, item_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """
    获取指定账号cookie_id和商品item_id的商品回复内容
    """
    try:
        user_id = current_user['user_id']
        # 校验cookie_id是否属于当前用户
        user_cookies = db_manager.get_all_cookies(user_id)
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        # 获取指定商品回复
        item_replies = db_manager.get_itemReplays_by_cookie(cookie_id)
        # 找对应item_id的回复
        item_reply = next((r for r in item_replies if r['item_id'] == item_id), None)

        if item_reply is None:
            raise HTTPException(status_code=404, detail="商品回复不存在")

        return item_reply

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品回复失败: {str(e)}")


# ------------------------- 数据库备份和恢复接口 -------------------------

@app.get('/admin/backup/download')
def download_database_backup(admin_user: Dict[str, Any] = Depends(require_permission('data_management'))):
    """下载数据库备份文件"""
    import os
    from fastapi.responses import FileResponse
    from datetime import datetime

    try:
        log_with_user('info', "请求下载数据库备份", admin_user)

        # 使用db_manager的实际数据库路径
        from db_manager import db_manager
        db_file_path = db_manager.db_path

        # 检查数据库文件是否存在
        if not os.path.exists(db_file_path):
            log_with_user('error', f"数据库文件不存在: {db_file_path}", admin_user)
            raise HTTPException(status_code=404, detail="数据库文件不存在")

        # 生成带时间戳的文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        download_filename = f"xianyu_backup_{timestamp}.db"

        log_with_user('info', f"开始下载数据库备份: {download_filename}", admin_user)

        return FileResponse(
            path=db_file_path,
            filename=download_filename,
            media_type='application/octet-stream'
        )

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"下载数据库备份失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.post('/admin/backup/upload')
async def upload_database_backup(admin_user: Dict[str, Any] = Depends(require_permission('import_backup')),
                                backup_file: UploadFile = File(...)):
    """上传并恢复数据库备份文件（管理员专用）"""
    import os
    import shutil
    import sqlite3
    from datetime import datetime

    try:
        log_with_user('info', f"开始上传数据库备份: {backup_file.filename}", admin_user)

        # 验证文件类型
        if not backup_file.filename.endswith('.db'):
            log_with_user('warning', f"无效的备份文件类型: {backup_file.filename}", admin_user)
            raise HTTPException(status_code=400, detail="只支持.db格式的数据库文件")

        # 验证文件大小（限制100MB）
        content = await backup_file.read()
        if len(content) > 100 * 1024 * 1024:  # 100MB
            log_with_user('warning', f"备份文件过大: {len(content)} bytes", admin_user)
            raise HTTPException(status_code=400, detail="备份文件大小不能超过100MB")

        # 验证是否为有效的SQLite数据库文件
        temp_file_path = f"temp_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"

        try:
            # 保存临时文件
            with open(temp_file_path, 'wb') as temp_file:
                temp_file.write(content)

            # 验证数据库文件完整性
            conn = sqlite3.connect(temp_file_path)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tables = cursor.fetchall()
            conn.close()

            # 检查是否包含必要的表
            table_names = [table[0] for table in tables]
            required_tables = ['users', 'cookies']  # 最基本的表

            missing_tables = [table for table in required_tables if table not in table_names]
            if missing_tables:
                log_with_user('warning', f"备份文件缺少必要的表: {missing_tables}", admin_user)
                raise HTTPException(status_code=400, detail=f"备份文件不完整，缺少表: {', '.join(missing_tables)}")

            log_with_user('info', f"备份文件验证通过，包含 {len(table_names)} 个表", admin_user)

        except sqlite3.Error as e:
            log_with_user('error', f"备份文件验证失败: {str(e)}", admin_user)
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
            raise HTTPException(status_code=400, detail="无效的数据库文件")

        # 备份当前数据库
        from db_manager import db_manager
        current_db_path = db_manager.db_path

        # 生成备份文件路径（与原数据库在同一目录）
        db_dir = os.path.dirname(current_db_path)
        backup_filename = f"xianyu_data_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        backup_current_path = os.path.join(db_dir, backup_filename)

        if os.path.exists(current_db_path):
            shutil.copy2(current_db_path, backup_current_path)
            log_with_user('info', f"当前数据库已备份为: {backup_current_path}", admin_user)

        # 关闭当前数据库连接
        if hasattr(db_manager, 'conn') and db_manager.conn:
            db_manager.conn.close()
            log_with_user('info', "已关闭当前数据库连接", admin_user)

        # 替换数据库文件
        shutil.move(temp_file_path, current_db_path)
        log_with_user('info', f"数据库文件已替换: {current_db_path}", admin_user)

        # 重新初始化数据库连接（使用原有的db_path）
        db_manager.__init__(db_manager.db_path)
        log_with_user('info', "数据库连接已重新初始化", admin_user)

        # 验证新数据库
        try:
            test_users = db_manager.get_all_users()
            log_with_user('info', f"数据库恢复成功，包含 {len(test_users)} 个用户", admin_user)
        except Exception as e:
            log_with_user('error', f"数据库恢复后验证失败: {str(e)}", admin_user)
            # 如果验证失败，尝试恢复原数据库
            if os.path.exists(backup_current_path):
                shutil.copy2(backup_current_path, current_db_path)
                db_manager.__init__()
                log_with_user('info', "已恢复原数据库", admin_user)
            raise HTTPException(status_code=500, detail="数据库恢复失败，已回滚到原数据库")

        return {
            "success": True,
            "message": "数据库恢复成功",
            "backup_file": backup_current_path,
            "user_count": len(test_users)
        }

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"上传数据库备份失败: {str(e)}", admin_user)
        # 清理临时文件
        if 'temp_file_path' in locals() and os.path.exists(temp_file_path):
            os.remove(temp_file_path)
        raise HTTPException(status_code=500, detail=str(e))

@app.get('/admin/backup/list')
def list_backup_files(admin_user: Dict[str, Any] = Depends(require_permission('data_management'))):
    """列出服务器上的备份文件（管理员专用）"""
    import os
    import glob
    from datetime import datetime

    try:
        log_with_user('info', "查询备份文件列表", admin_user)

        # 查找备份文件
        backup_files = glob.glob("xianyu_data_backup_*.db")

        backup_list = []
        for file_path in backup_files:
            try:
                stat = os.stat(file_path)
                backup_list.append({
                    'filename': os.path.basename(file_path),
                    'size': stat.st_size,
                    'size_mb': round(stat.st_size / (1024 * 1024), 2),
                    'created_time': datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S'),
                    'modified_time': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                })
            except Exception as e:
                log_with_user('warning', f"读取备份文件信息失败: {file_path} - {str(e)}", admin_user)

        # 按修改时间倒序排列
        backup_list.sort(key=lambda x: x['modified_time'], reverse=True)

        log_with_user('info', f"找到 {len(backup_list)} 个备份文件", admin_user)

        return {
            "backups": backup_list,
            "total": len(backup_list)
        }

    except Exception as e:
        log_with_user('error', f"查询备份文件列表失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 数据管理接口 -------------------------

@app.get('/admin/orders')
def get_orders_with_item_title(admin_user: Dict[str, Any] = Depends(require_permission('data_management'))):
    """获取订单列表（含商品标题）"""
    from db_manager import db_manager
    try:
        log_with_user('info', "查询订单列表（含商品标题）", admin_user)

        with db_manager.lock:
            cursor = db_manager.conn.cursor()
            cursor.execute('''
                SELECT o.order_id, o.item_id, o.buyer_id, o.spec_name, o.spec_value,
                       o.quantity, o.amount, o.order_status, o.cookie_id,
                       o.created_at, o.updated_at, i.item_title,
                       o.delivery_status, o.buyer_name
                FROM orders o
                LEFT JOIN item_info i ON o.cookie_id = i.cookie_id AND o.item_id = i.item_id
                ORDER BY o.created_at DESC
            ''')

            data = []
            for row in cursor.fetchall():
                data.append({
                    'order_id': row[0],
                    'item_id': row[1],
                    'buyer_id': row[2],
                    'spec_name': row[3],
                    'spec_value': row[4],
                    'quantity': row[5],
                    'amount': row[6],
                    'order_status': row[7],
                    'cookie_id': row[8],
                    'created_at': row[9],
                    'updated_at': row[10],
                    'item_title': row[11],
                    'delivery_status': row[12] if row[12] else 'pending',
                    'buyer_name': row[13] if row[13] else ''
                })

        log_with_user('info', f"订单列表查询成功，共 {len(data)} 条记录", admin_user)

        return {
            "success": True,
            "data": data,
            "count": len(data)
        }

    except Exception as e:
        log_with_user('error', f"查询订单列表失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))


@app.get('/admin/data/{table_name}')
def get_table_data(table_name: str, admin_user: Dict[str, Any] = Depends(require_permission('data_management'))):
    """获取指定表的所有数据"""
    from db_manager import db_manager
    try:
        log_with_user('info', f"查询表数据: {table_name}", admin_user)

        # 验证表名安全性
        allowed_tables = [
            'users', 'cookies', 'cookie_status', 'keywords', 'default_replies', 'default_reply_records',
            'ai_reply_settings', 'ai_conversations', 'ai_item_cache', 'item_info',
            'message_notifications', 'cards', 'delivery_rules', 'notification_channels',
            'user_settings', 'system_settings', 'email_verifications', 'captcha_codes', 'orders', "item_replay"
        ]

        if table_name not in allowed_tables:
            log_with_user('warning', f"尝试访问不允许的表: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="不允许访问该表")

        # 获取表数据
        data, columns = db_manager.get_table_data(table_name)

        log_with_user('info', f"表 {table_name} 查询成功，共 {len(data)} 条记录", admin_user)

        return {
            "success": True,
            "data": data,
            "columns": columns,
            "count": len(data)
        }

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"查询表数据失败: {table_name} - {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.delete('/admin/data/{table_name}/{record_id}')
def delete_table_record(table_name: str, record_id: str, admin_user: Dict[str, Any] = Depends(require_permission('data_management'))):
    """删除指定表的指定记录"""
    from db_manager import db_manager
    try:
        log_with_user('info', f"删除表记录: {table_name}.{record_id}", admin_user)

        # 验证表名安全性
        allowed_tables = [
            'users', 'cookies', 'cookie_status', 'keywords', 'default_replies', 'default_reply_records',
            'ai_reply_settings', 'ai_conversations', 'ai_item_cache', 'item_info',
            'message_notifications', 'cards', 'delivery_rules', 'notification_channels',
            'user_settings', 'system_settings', 'email_verifications', 'captcha_codes', 'orders','item_replay'
        ]

        if table_name not in allowed_tables:
            log_with_user('warning', f"尝试删除不允许的表记录: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="不允许操作该表")

        # 特殊保护：不能删除管理员用户
        if table_name == 'users' and record_id == str(admin_user['user_id']):
            log_with_user('warning', "尝试删除管理员自己", admin_user)
            raise HTTPException(status_code=400, detail="不能删除管理员自己")

        # 删除记录
        success = db_manager.delete_table_record(table_name, record_id)

        if success:
            log_with_user('info', f"表记录删除成功: {table_name}.{record_id}", admin_user)
            return {"success": True, "message": "删除成功"}
        else:
            log_with_user('warning', f"表记录删除失败: {table_name}.{record_id}", admin_user)
            raise HTTPException(status_code=400, detail="删除失败，记录可能不存在")

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"删除表记录异常: {table_name}.{record_id} - {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@app.delete('/admin/data/{table_name}')
def clear_table_data(table_name: str, admin_user: Dict[str, Any] = Depends(require_permission('data_management'))):
    """清空指定表的所有数据"""
    from db_manager import db_manager
    try:
        log_with_user('info', f"清空表数据: {table_name}", admin_user)

        # 验证表名安全性
        allowed_tables = [
            'cookies', 'cookie_status', 'keywords', 'default_replies', 'default_reply_records',
            'ai_reply_settings', 'ai_conversations', 'ai_item_cache', 'item_info',
            'message_notifications', 'cards', 'delivery_rules', 'notification_channels',
            'user_settings', 'system_settings', 'email_verifications', 'captcha_codes', 'orders', "item_replay"
        ]

        # 不允许清空用户表
        if table_name == 'users':
            log_with_user('warning', "尝试清空用户表", admin_user)
            raise HTTPException(status_code=400, detail="不允许清空用户表")

        if table_name not in allowed_tables:
            log_with_user('warning', f"尝试清空不允许的表: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="不允许清空该表")

        # 清空表数据
        success = db_manager.clear_table_data(table_name)

        if success:
            log_with_user('info', f"表数据清空成功: {table_name}", admin_user)
            return {"success": True, "message": "清空成功"}
        else:
            log_with_user('warning', f"表数据清空失败: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="清空失败")

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"清空表数据异常: {table_name} - {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))


# 订单重新发货API
@app.post("/orders/{order_id}/retry-delivery")
async def retry_order_delivery(order_id: str, admin_user: Dict[str, Any] = Depends(require_permission('data_management'))):
    """重新发货 - 对失败的订单重新执行自动发货"""
    try:
        from db_manager import db_manager

        log_with_user('info', f"尝试重新发货: 订单 {order_id}", admin_user)

        # 获取订单信息
        order = db_manager.get_order_by_id(order_id)
        if not order:
            raise HTTPException(status_code=404, detail="订单不存在")

        cookie_id = order.get('cookie_id')
        item_id = order.get('item_id')
        buyer_id = order.get('buyer_id')

        if not cookie_id or not item_id:
            raise HTTPException(status_code=400, detail="订单缺少必要的商品或账号信息")

        # 查找对应的 XianyuLive 实例
        if cookie_manager.manager is None:
            raise HTTPException(status_code=400, detail="Cookie管理器未初始化")

        # 找到商品真正的卖家账号
        seller_cookie_id = cookie_id
        try:
            cursor = db_manager.conn.cursor()
            cursor.execute('SELECT cookie_id FROM item_info WHERE item_id = ?', (item_id,))
            row = cursor.fetchone()
            if row and row[0]:
                seller_cookie_id = row[0]
                log_with_user('info', f"重新发货: 找到商品卖家账号 {seller_cookie_id}", admin_user)
        except Exception:
            pass

        xianyu_instance = cookie_manager.manager.get_xianyu(seller_cookie_id)
        if not xianyu_instance:
            # 尝试原始cookie_id
            xianyu_instance = cookie_manager.manager.get_xianyu(cookie_id)
            if not xianyu_instance:
                raise HTTPException(status_code=400, detail=f"账号 {seller_cookie_id} 未在线，无法重新发货")

        # 重置订单状态为处理中
        db_manager.update_order_delivery_status(order_id, 'pending')

        # 先尝试获取商品标题（用于匹配发货规则）
        item_title = None
        try:
            # 1. 从所有账号的商品信息中查找
            cursor = db_manager.conn.cursor()
            cursor.execute('SELECT item_title FROM item_info WHERE item_id = ?', (item_id,))
            row = cursor.fetchone()
            if row and row[0]:
                item_title = row[0]
        except Exception:
            pass

        if not item_title:
            try:
                # 2. 通过API获取（用当前账号）
                api_result = await xianyu_instance.get_item_info(item_id)
                if api_result and 'data' in api_result:
                    data = api_result['data']
                    item_data = data.get('itemDO', {})
                    share_data = item_data.get('shareData', {})
                    share_info_str = share_data.get('shareInfoJsonString', '')
                    if share_info_str:
                        import json
                        share_info = json.loads(share_info_str)
                        item_title = share_info.get('contentParams', {}).get('mainParams', {}).get('content', '')
                    if not item_title:
                        item_title = item_data.get('title', '')
            except Exception:
                pass

        # 3. 尝试从发货规则反向查找（用订单的item_id匹配）
        if not item_title:
            try:
                cursor = db_manager.conn.cursor()
                cursor.execute('SELECT keyword FROM delivery_rules WHERE enabled = 1')
                rules = cursor.fetchall()
                # 返回所有规则关键字供前端选择
                rule_keywords = [r[0] for r in rules]
                log_with_user('warning', f"无法获取商品标题，但找到发货规则: {rule_keywords}", admin_user)
            except Exception:
                pass

        if not item_title:
            item_title = "未知商品"

        log_with_user('info', f"重新发货: 订单 {order_id}, 商品ID={item_id}, 商品标题={item_title[:50]}", admin_user)

        # 调用自动发货方法
        delivery_content = await xianyu_instance._auto_delivery(item_id, item_title, order_id, buyer_id)

        if delivery_content:
            # 发货成功，更新状态
            db_manager.update_order_delivery_status(order_id, 'delivered')

            # 发送发货消息（用卖家的聊天ID格式）
            chat_id = f"{item_id}_{buyer_id}"
            ws = xianyu_instance.ws
            if not ws:
                log_with_user('warning', f"重新发货: WebSocket未连接，无法发送消息 {order_id}", admin_user)
                return {"success": True, "message": "发货内容获取成功，但WebSocket未连接，消息未发送"}

            # 获取商品的session_id用于构造正确的chat_id
            try:
                cursor = db_manager.conn.cursor()
                cursor.execute("SELECT value FROM cookies WHERE id = ?", (seller_cookie_id,))
                cookie_row = cursor.fetchone()
                if cookie_row:
                    from cookie_manager import manager as cm
                    if cm and seller_cookie_id in cm.keywords:
                        pass
            except Exception:
                pass

            if delivery_content.startswith("__IMAGE_SEND__"):
                image_data = delivery_content.replace("__IMAGE_SEND__", "")
                if "|" in image_data:
                    card_id_str, image_url = image_data.split("|", 1)
                    try:
                        card_id = int(card_id_str)
                    except ValueError:
                        card_id = None
                    await xianyu_instance.send_image_msg(ws, chat_id, buyer_id, image_url, card_id=card_id)
                else:
                    await xianyu_instance.send_image_msg(ws, chat_id, buyer_id, image_data)
            else:
                await xianyu_instance.send_msg(ws, chat_id, buyer_id, delivery_content)

            log_with_user('info', f"重新发货成功: 订单 {order_id}", admin_user)
            return {"success": True, "message": "重新发货成功"}
        else:
            # 发货失败
            db_manager.update_order_delivery_status(order_id, 'failed')
            log_with_user('warning', f"重新发货失败: 订单 {order_id}", admin_user)
            raise HTTPException(status_code=400, detail="重新发货失败，未找到匹配的发货规则或获取发货内容失败")

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"重新发货异常: {order_id} - {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))


# 商品多规格管理API
@app.put("/items/{cookie_id}/{item_id}/multi-spec")
def update_item_multi_spec(cookie_id: str, item_id: str, spec_data: dict, _: None = Depends(require_auth)):
    """更新商品的多规格状态"""
    try:
        from db_manager import db_manager

        is_multi_spec = spec_data.get('is_multi_spec', False)

        success = db_manager.update_item_multi_spec_status(cookie_id, item_id, is_multi_spec)

        if success:
            return {"message": f"商品多规格状态已{'开启' if is_multi_spec else '关闭'}"}
        else:
            raise HTTPException(status_code=404, detail="商品不存在")

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# 商品多数量发货管理API
@app.put("/items/{cookie_id}/{item_id}/multi-quantity-delivery")
def update_item_multi_quantity_delivery(cookie_id: str, item_id: str, delivery_data: dict, _: None = Depends(require_auth)):
    """更新商品的多数量发货状态"""
    try:
        from db_manager import db_manager

        multi_quantity_delivery = delivery_data.get('multi_quantity_delivery', False)

        success = db_manager.update_item_multi_quantity_delivery_status(cookie_id, item_id, multi_quantity_delivery)

        if success:
            return {"message": f"商品多数量发货状态已{'开启' if multi_quantity_delivery else '关闭'}"}
        else:
            raise HTTPException(status_code=404, detail="商品不存在")

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# 移除自动启动，由Start.py或手动启动
# if __name__ == "__main__":
#     uvicorn.run(app, host="0.0.0.0", port=8080)


# ========================= 阿奇索账户管理API =========================

class AgisoAccountRequest(BaseModel):
    name: str
    cookie: str
    authorization: str

class AgisoAccountUpdateRequest(BaseModel):
    name: Optional[str] = None
    cookie: Optional[str] = None
    authorization: Optional[str] = None

class ResellRequest(BaseModel):
    agiso_account_id: int
    source_item: Dict[str, Any]
    resell_price: str
    stock: int = 1
    description: str = ""


@app.get("/agiso/accounts")
async def get_agiso_accounts(
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """获取当前用户的所有阿奇索账户"""
    if not current_user:
        raise HTTPException(status_code=401, detail="未登录")
    
    try:
        from db_manager import db_manager
        accounts = db_manager.get_agiso_accounts(current_user["user_id"])
        # 隐藏敏感信息
        for account in accounts:
            if account.get("cookie"):
                account["cookie"] = account["cookie"][:20] + "..." if len(account["cookie"]) > 20 else "***"
            if account.get("authorization"):
                account["authorization"] = account["authorization"][:20] + "..." if len(account["authorization"]) > 20 else "***"
        return {"success": True, "data": accounts}
    except Exception as e:
        logger.error(f"获取阿奇索账户列表失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/agiso/accounts")
async def add_agiso_account(
    request: AgisoAccountRequest,
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """添加阿奇索账户"""
    if not current_user:
        raise HTTPException(status_code=401, detail="未登录")
    
    try:
        from db_manager import db_manager
        
        # 尝试验证账户有效性（可选，即使失败也允许添加）
        verify_message = ""
        try:
            from utils.agiso_client import AgisoClient
            async with AgisoClient(request.cookie, request.authorization) as client:
                verify_result = await client.verify_account()
                if not verify_result.get("success"):
                    verify_message = f"（注意：{verify_result.get('message')}，但账户已添加，后续使用时可能需要重新配置）"
        except Exception as e:
            logger.warning(f"验证阿奇索账户时出错: {e}")
            verify_message = "（账户已添加，验证功能暂时不可用）"
        
        # 添加账户
        account_id = db_manager.add_agiso_account(
            user_id=current_user["user_id"],
            name=request.name,
            cookie=request.cookie,
            authorization=request.authorization
        )
        
        if account_id == -1:
            return {"success": False, "message": "添加账户失败"}
        
        message = f"添加成功{verify_message}" if verify_message else "添加成功"
        return {"success": True, "message": message, "account_id": account_id}
    except Exception as e:
        logger.error(f"添加阿奇索账户失败: {e}")
        return {"success": False, "message": str(e)}


@app.put("/agiso/accounts/{account_id}")
async def update_agiso_account(
    account_id: int,
    request: AgisoAccountUpdateRequest,
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """更新阿奇索账户"""
    if not current_user:
        raise HTTPException(status_code=401, detail="未登录")
    
    try:
        from db_manager import db_manager
        
        # 验证账户归属
        account = db_manager.get_agiso_account_by_id(account_id)
        if not account or account["user_id"] != current_user["user_id"]:
            raise HTTPException(status_code=404, detail="账户不存在")
        
        # 更新账户
        success = db_manager.update_agiso_account(
            account_id=account_id,
            name=request.name,
            cookie=request.cookie,
            authorization=request.authorization
        )
        
        if success:
            return {"success": True, "message": "更新成功"}
        else:
            return {"success": False, "message": "更新失败"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新阿奇索账户失败: {e}")
        return {"success": False, "message": str(e)}


@app.delete("/agiso/accounts/{account_id}")
async def delete_agiso_account(
    account_id: int,
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """删除阿奇索账户"""
    if not current_user:
        raise HTTPException(status_code=401, detail="未登录")
    
    try:
        from db_manager import db_manager
        
        # 验证账户归属
        account = db_manager.get_agiso_account_by_id(account_id)
        if not account or account["user_id"] != current_user["user_id"]:
            raise HTTPException(status_code=404, detail="账户不存在")
        
        # 删除账户
        success = db_manager.delete_agiso_account(account_id)
        
        if success:
            return {"success": True, "message": "删除成功"}
        else:
            return {"success": False, "message": "删除失败"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除阿奇索账户失败: {e}")
        return {"success": False, "message": str(e)}


@app.post("/agiso/verify")
async def verify_agiso_account(
    account_id: int,
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """验证阿奇索账户有效性"""
    if not current_user:
        raise HTTPException(status_code=401, detail="未登录")
    
    try:
        from db_manager import db_manager
        
        # 获取账户信息
        account = db_manager.get_agiso_account_by_id(account_id)
        if not account or account["user_id"] != current_user["user_id"]:
            raise HTTPException(status_code=404, detail="账户不存在")
        
        # 验证账户
        from utils.agiso_client import AgisoClient
        async with AgisoClient(account["cookie"], account["authorization"]) as client:
            result = await client.verify_account()
            return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"验证阿奇索账户失败: {e}")
        return {"success": False, "message": str(e)}


# ========================= 转卖操作API =========================

@app.post("/resell/publish")
async def resell_publish(
    request: ResellRequest,
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """发布转卖商品"""
    if not current_user:
        raise HTTPException(status_code=401, detail="未登录")
    
    try:
        from db_manager import db_manager
        from utils.item_publisher import item_publisher
        
        # 验证阿奇索账户归属
        account = db_manager.get_agiso_account_by_id(request.agiso_account_id)
        if not account or account["user_id"] != current_user["user_id"]:
            return {"success": False, "message": "阿奇索账户不存在"}
        
        # 执行转卖
        result = await item_publisher.resell_item(
            source_item=request.source_item,
            agiso_account=account,
            resell_price=request.resell_price,
            stock=request.stock,
            description=request.description,
            user_id=current_user["user_id"]
        )
        
        return result
    except Exception as e:
        logger.error(f"转卖发布失败: {e}")
        return {"success": False, "message": str(e)}


@app.get("/resell/records")
async def get_resell_records(
    limit: int = 50,
    offset: int = 0,
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """获取转卖记录"""
    if not current_user:
        raise HTTPException(status_code=401, detail="未登录")
    
    try:
        from db_manager import db_manager
        records = db_manager.get_resell_records(
            user_id=current_user["user_id"],
            limit=limit,
            offset=offset
        )
        return {"success": True, "data": records}
    except Exception as e:
        logger.error(f"获取转卖记录失败: {e}")
        return {"success": False, "data": []}


@app.get("/resell/records/{record_id}")
async def get_resell_record_detail(
    record_id: int,
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """获取单条转卖记录详情"""
    if not current_user:
        raise HTTPException(status_code=401, detail="未登录")
    
    try:
        from db_manager import db_manager
        record = db_manager.get_resell_record_by_id(record_id)
        
        if not record or record.get("user_id") != current_user["user_id"]:
            raise HTTPException(status_code=404, detail="记录不存在")
        
        return {"success": True, "data": record}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取转卖记录详情失败: {e}")
        return {"success": False, "message": str(e)}


@app.delete("/resell/records/{record_id}")
async def delete_resell_record(
    record_id: int,
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """删除转卖记录"""
    if not current_user:
        raise HTTPException(status_code=401, detail="未登录")
    
    try:
        from db_manager import db_manager
        
        # 验证记录归属
        record = db_manager.get_resell_record_by_id(record_id)
        if not record or record.get("user_id") != current_user["user_id"]:
            raise HTTPException(status_code=404, detail="记录不存在")
        
        # 删除记录
        with db_manager.lock:
            cursor = db_manager.conn.cursor()
            db_manager._execute_sql(cursor, "DELETE FROM resell_records WHERE id = ?", (record_id,))
            db_manager.conn.commit()
        
        return {"success": True, "message": "删除成功"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除转卖记录失败: {e}")
        return {"success": False, "message": str(e)}